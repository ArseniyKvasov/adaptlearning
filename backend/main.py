from __future__ import annotations

import json
import math
import os
import random
import sqlite3
import subprocess
import tempfile
from io import BytesIO
import uuid
import asyncio
import hashlib
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional
from zipfile import ZIP_DEFLATED, ZipFile
from xml.sax.saxutils import escape as xml_escape

from dotenv import load_dotenv
from fastapi import BackgroundTasks, FastAPI, File, HTTPException, Request, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from backend.ml_service import MLServiceClient, MLServiceError

BASE_DIR = Path(__file__).resolve().parent.parent
load_dotenv(BASE_DIR / ".env")
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = Path(os.getenv("DB_PATH", str(DATA_DIR / "app.db")))
PUBLIC_DIR = BASE_DIR / "public"

ML_URL = os.getenv("ML_URL", "https://ml.fastclass.ru")
ML_API_KEY = os.getenv("ML_API_KEY", "")
AUDIO_SAMPLE_RATE = int(os.getenv("AUDIO_SAMPLE_RATE", "16000"))
AUDIO_CHUNK_SECONDS = int(os.getenv("AUDIO_CHUNK_SECONDS", "480"))
AUDIO_CHUNK_MIN_SECONDS = int(os.getenv("AUDIO_CHUNK_MIN_SECONDS", "300"))
AUDIO_CHUNK_MAX_SECONDS = int(os.getenv("AUDIO_CHUNK_MAX_SECONDS", "600"))
AUDIO_CHUNK_OVERLAP_SECONDS = int(os.getenv("AUDIO_CHUNK_OVERLAP_SECONDS", "0"))
AUDIO_SILENCE_NOISE_DB = os.getenv("AUDIO_SILENCE_NOISE_DB", "-35dB")
AUDIO_SILENCE_MIN_SECONDS = float(os.getenv("AUDIO_SILENCE_MIN_SECONDS", "0.7"))
SPEECH_ANALYSIS_GROUP_TARGET_SECONDS = int(os.getenv("SPEECH_ANALYSIS_GROUP_TARGET_SECONDS", str(7 * 60)))
SPEECH_ANALYSIS_GROUP_MIN_SECONDS = int(os.getenv("SPEECH_ANALYSIS_GROUP_MIN_SECONDS", str(6 * 60)))
SPEECH_ANALYSIS_GROUP_MAX_SECONDS = int(os.getenv("SPEECH_ANALYSIS_GROUP_MAX_SECONDS", str(8 * 60)))
SPEECH_ANALYSIS_GROUP_OVERLAP_PHRASES = int(os.getenv("SPEECH_ANALYSIS_GROUP_OVERLAP_PHRASES", "3"))
TRANSCRIBE_BATCH_SIZE = max(1, int(os.getenv("TRANSCRIBE_BATCH_SIZE", "2")))

app = FastAPI()
app.add_middleware(SessionMiddleware, secret_key=os.getenv("SESSION_SECRET", "dev-secret-change-me"), same_site="lax", https_only=False, max_age=60 * 60 * 24 * 30)
WS_CONNECTIONS = {}


def db_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    conn = db_conn()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
          id TEXT PRIMARY KEY,
          created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS generations (
          id TEXT PRIMARY KEY,
          user_id TEXT NOT NULL,
          creator_id TEXT NOT NULL DEFAULT '',
          file_name TEXT NOT NULL,
          status TEXT NOT NULL,
          progress_percent REAL NOT NULL DEFAULT 0,
          created_at TEXT NOT NULL,
          transcript_json TEXT NOT NULL,
          mini_summary_json TEXT NOT NULL,
          summary_json TEXT NOT NULL,
          quiz_json TEXT NOT NULL,
          practice_json TEXT NOT NULL,
          analytics_json TEXT NOT NULL,
          FOREIGN KEY (user_id) REFERENCES users(id)
        )
        """
    )
    cols = [r[1] for r in conn.execute("PRAGMA table_info(generations)").fetchall()]
    if "creator_id" not in cols:
        conn.execute("ALTER TABLE generations ADD COLUMN creator_id TEXT NOT NULL DEFAULT ''")
        conn.execute(
            """
            UPDATE generations
            SET creator_id = CASE
                WHEN creator_id IS NULL OR trim(creator_id) = '' THEN user_id
                ELSE creator_id
            END
            WHERE creator_id IS NULL OR trim(creator_id) = ''
            """
        )
    cols = [r[1] for r in conn.execute("PRAGMA table_info(generations)").fetchall()]
    if "error_message" not in cols:
        conn.execute("ALTER TABLE generations ADD COLUMN error_message TEXT NOT NULL DEFAULT ''")
    cols = [r[1] for r in conn.execute("PRAGMA table_info(generations)").fetchall()]
    if "progress_percent" not in cols:
        conn.execute("ALTER TABLE generations ADD COLUMN progress_percent REAL NOT NULL DEFAULT 0")
    cols = [r[1] for r in conn.execute("PRAGMA table_info(generations)").fetchall()]
    if "mini_summary_json" not in cols:
        conn.execute("ALTER TABLE generations ADD COLUMN mini_summary_json TEXT NOT NULL DEFAULT '[]'")
    cols = [r[1] for r in conn.execute("PRAGMA table_info(generations)").fetchall()]
    if "practice_json" not in cols:
        conn.execute("ALTER TABLE generations ADD COLUMN practice_json TEXT NOT NULL DEFAULT '{}'")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS transcript_cache (
          content_hash TEXT PRIMARY KEY,
          transcript_json TEXT NOT NULL,
          created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS student_attempts (
          id TEXT PRIMARY KEY,
          generation_id TEXT NOT NULL,
          user_id TEXT NOT NULL,
          created_at TEXT NOT NULL,
          answers_json TEXT NOT NULL,
          results_json TEXT NOT NULL,
          mastery_json TEXT NOT NULL,
          recommendation TEXT NOT NULL,
          subtopic_to_revise TEXT NOT NULL,
          FOREIGN KEY (generation_id) REFERENCES generations(id),
          FOREIGN KEY (user_id) REFERENCES users(id)
        )
        """
    )
    cols = [r[1] for r in conn.execute("PRAGMA table_info(student_attempts)").fetchall()]
    if "user_id" not in cols:
        conn.execute("ALTER TABLE student_attempts ADD COLUMN user_id TEXT NOT NULL DEFAULT ''")
    conn.execute(
        """
        UPDATE student_attempts
        SET user_id = CASE
            WHEN user_id IS NULL OR trim(user_id) = '' THEN 'legacy_' || id
            ELSE user_id
        END
        WHERE user_id IS NULL OR trim(user_id) = ''
        """
    )
    cleanup_student_attempt_duplicates(conn)
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_student_attempts_generation_user ON student_attempts(generation_id, user_id)")
    conn.commit()
    conn.close()


@app.on_event("startup")
async def startup() -> None:
    init_db()


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def ensure_guest_user(request: Request) -> str:
    user_id = request.session.get("user_id")
    if user_id:
        return user_id

    user_id = f"guest_{uuid.uuid4().hex[:14]}"
    conn = db_conn()
    user_cols = {row["name"] for row in conn.execute("PRAGMA table_info(users)").fetchall()}
    if "role" in user_cols:
        conn.execute("INSERT INTO users (id, role, created_at) VALUES (?, ?, ?)", (user_id, "guest", now_iso()))
    else:
        conn.execute("INSERT INTO users (id, created_at) VALUES (?, ?)", (user_id, now_iso()))
    conn.commit()
    conn.close()
    request.session["user_id"] = user_id
    return user_id


def row_to_generation(row: sqlite3.Row) -> dict[str, Any]:
    creator_id = row["creator_id"] if "creator_id" in row.keys() else row["user_id"]
    practice_raw = {}
    if "practice_json" in row.keys():
        try:
            practice_raw = json.loads(row["practice_json"]) if row["practice_json"] else {}
        except json.JSONDecodeError:
            practice_raw = {}
    return {
        "id": row["id"],
        "user_id": row["user_id"],
        "creator_id": creator_id,
        "file_name": row["file_name"],
        "status": row["status"],
        "progress_percent": float(row["progress_percent"]) if "progress_percent" in row.keys() and row["progress_percent"] is not None else 0,
        "created_at": row["created_at"],
        "transcript": json.loads(row["transcript_json"]),
        "mini_summary": json.loads(row["mini_summary_json"]) if "mini_summary_json" in row.keys() else [],
        "summary": json.loads(row["summary_json"]),
        "quiz": json.loads(row["quiz_json"]),
        "practice": practice_raw if isinstance(practice_raw, dict) else {},
        "analytics": json.loads(row["analytics_json"]),
        "error_message": row["error_message"] if "error_message" in row.keys() else "",
    }


def content_hash_for_bytes(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


def sanitize_uploaded_filename(file_name: str) -> str:
    raw_name = str(file_name or "").strip()
    if not raw_name:
        return "media"

    path = Path(raw_name)
    suffix = path.suffix.lower()
    if suffix and not re.fullmatch(r"\.[a-z0-9]{1,8}", suffix):
        suffix = ""

    stem_source = path.name[:-len(suffix)] if suffix else path.stem
    stem = re.sub(r"[^0-9A-Za-zА-Яа-яЁё_-]+", "_", stem_source)
    stem = re.sub(r"_+", "_", stem).strip("_-")
    if not stem:
        stem = "media"
    stem = stem[:50]
    if not stem:
        stem = "media"
    return f"{stem}{suffix}"


def get_cached_transcript(content_hash: str) -> Optional[list[dict[str, Any]]]:
    conn = db_conn()
    row = conn.execute("SELECT transcript_json FROM transcript_cache WHERE content_hash = ?", (content_hash,)).fetchone()
    conn.close()
    if not row:
        return None
    try:
        transcript = json.loads(row["transcript_json"])
    except json.JSONDecodeError:
        return None
    return transcript if isinstance(transcript, list) else None


def store_cached_transcript(content_hash: str, transcript: list[dict[str, Any]]) -> None:
    conn = db_conn()
    conn.execute(
        """
        INSERT INTO transcript_cache (content_hash, transcript_json, created_at)
        VALUES (?, ?, ?)
        ON CONFLICT(content_hash) DO UPDATE SET
          transcript_json = excluded.transcript_json,
          created_at = excluded.created_at
        """,
        (content_hash, json.dumps(transcript, ensure_ascii=False), now_iso()),
    )
    conn.commit()
    conn.close()


def cleanup_student_attempt_duplicates(conn: sqlite3.Connection) -> None:
    rows = conn.execute(
        """
        SELECT rowid, generation_id, user_id, created_at, id
        FROM student_attempts
        ORDER BY generation_id, user_id, datetime(created_at) DESC, id DESC
        """
    ).fetchall()
    seen_pairs: set[tuple[str, str]] = set()
    for row in rows:
        pair = (str(row["generation_id"] or ""), str(row["user_id"] or ""))
        if pair in seen_pairs:
            conn.execute("DELETE FROM student_attempts WHERE rowid = ?", (row["rowid"],))
            continue
        seen_pairs.add(pair)


def get_generation(generation_id: str) -> Optional[dict[str, Any]]:
    conn = db_conn()
    row = conn.execute("SELECT * FROM generations WHERE id = ?", (generation_id,)).fetchone()
    conn.close()
    return row_to_generation(row) if row else None


def default_practice_state() -> dict[str, Any]:
    return {
        "status": "idle",
        "stage": "",
        "weak_subtopics": [],
        "current_weak_subtopics": [],
        "pending_weak_subtopics": [],
        "mastery": {},
        "mastery_order": [],
        "practice_round": 0,
        "round_submitted": False,
        "practice_completed": False,
        "request": {},
        "summary": [],
        "quiz": [],
        "error_message": "",
        "stale_reason": "",
        "updated_at": "",
    }


def normalize_practice_state(raw_state: Any) -> dict[str, Any]:
    state = default_practice_state()
    if isinstance(raw_state, dict):
        try:
            practice_round = int(raw_state.get("practice_round") or state["practice_round"] or 0)
        except (TypeError, ValueError):
            practice_round = state["practice_round"]
        state.update({
            "status": str(raw_state.get("status") or state["status"]),
            "stage": str(raw_state.get("stage") or state["stage"]),
            "error_message": str(raw_state.get("error_message") or state["error_message"]),
            "stale_reason": str(raw_state.get("stale_reason") or state["stale_reason"]),
            "updated_at": str(raw_state.get("updated_at") or state["updated_at"]),
            "practice_round": practice_round,
            "round_submitted": bool(raw_state.get("round_submitted", state["round_submitted"])),
            "practice_completed": bool(raw_state.get("practice_completed", state["practice_completed"])),
        })
        weak_subtopics = raw_state.get("weak_subtopics")
        if isinstance(weak_subtopics, list):
            state["weak_subtopics"] = [str(item).strip() for item in weak_subtopics if str(item).strip()]
        current_weak_subtopics = raw_state.get("current_weak_subtopics")
        if isinstance(current_weak_subtopics, list):
            state["current_weak_subtopics"] = [str(item).strip() for item in current_weak_subtopics if str(item).strip()]
        pending_weak_subtopics = raw_state.get("pending_weak_subtopics")
        if isinstance(pending_weak_subtopics, list):
            state["pending_weak_subtopics"] = [str(item).strip() for item in pending_weak_subtopics if str(item).strip()]
        mastery = raw_state.get("mastery")
        if isinstance(mastery, dict):
            state["mastery"] = normalize_mastery_map(mastery)
        elif isinstance(mastery, list):
            state["mastery"] = normalize_mastery_map(mastery)
        mastery_order = raw_state.get("mastery_order")
        if isinstance(mastery_order, list):
            state["mastery_order"] = [str(item).strip() for item in mastery_order if str(item).strip()]
        request = raw_state.get("request")
        if isinstance(request, dict):
            state["request"] = request
        summary = raw_state.get("summary")
        if isinstance(summary, list):
            state["summary"] = summary
        quiz = raw_state.get("quiz")
        if isinstance(quiz, list):
            state["quiz"] = quiz
    return state


def normalize_mastery_map(raw_mastery: Any) -> dict[str, int]:
    mastery: dict[str, int] = {}
    items: list[tuple[str, Any]] = []
    if isinstance(raw_mastery, dict):
        items = [(str(key), value) for key, value in raw_mastery.items()]
    elif isinstance(raw_mastery, list):
        for item in raw_mastery:
            if not isinstance(item, dict):
                continue
            subtopic = str(item.get("subtopic") or "").strip()
            if not subtopic:
                continue
            items.append((subtopic, item.get("percent", 0)))
    for subtopic_raw, percent_raw in items:
        subtopic = str(subtopic_raw or "").strip()
        if not subtopic:
            continue
        try:
            percent = int(percent_raw or 0)
        except (TypeError, ValueError):
            percent = 0
        mastery[subtopic] = max(0, min(100, percent))
    return mastery


def practice_mastery_order(practice: dict[str, Any], fallback_order: list[str] | None = None) -> list[str]:
    order: list[str] = []
    seen: set[str] = set()
    raw_order = practice.get("mastery_order")
    if isinstance(raw_order, list):
        for item in raw_order:
            subtopic = str(item or "").strip()
            if not subtopic or subtopic in seen:
                continue
            order.append(subtopic)
            seen.add(subtopic)
    if fallback_order:
        for item in fallback_order:
            subtopic = str(item or "").strip()
            if not subtopic or subtopic in seen:
                continue
            order.append(subtopic)
            seen.add(subtopic)
    mastery = normalize_mastery_map(practice.get("mastery", {}))
    for subtopic in mastery.keys():
        if subtopic not in seen:
            order.append(subtopic)
            seen.add(subtopic)
    return order


def practice_low_topics(mastery: dict[str, int], order: list[str]) -> list[dict[str, Any]]:
    order_map = {subtopic: idx for idx, subtopic in enumerate(order)}
    low_topics = [
        {"subtopic": subtopic, "percent": int(percent or 0)}
        for subtopic, percent in mastery.items()
        if int(percent or 0) < 80
    ]
    low_topics.sort(
        key=lambda item: (
            int(item.get("percent", 0) or 0),
            order_map.get(str(item.get("subtopic") or ""), len(order_map)),
            str(item.get("subtopic") or "").casefold(),
        )
    )
    return low_topics


def practice_round_topics(mastery: dict[str, int], order: list[str], limit: int = 2) -> tuple[list[str], list[str]]:
    low_topics = practice_low_topics(mastery, order)
    selected = [str(item.get("subtopic") or "").strip() for item in low_topics[:limit] if str(item.get("subtopic") or "").strip()]
    pending = [str(item.get("subtopic") or "").strip() for item in low_topics[limit:] if str(item.get("subtopic") or "").strip()]
    return selected, pending


def merge_practice_mastery(practice: dict[str, Any], round_mastery: list[dict[str, Any]], fallback_order: list[str] | None = None) -> dict[str, Any]:
    mastery = normalize_mastery_map(practice.get("mastery", {}))
    for item in round_mastery:
        if not isinstance(item, dict):
            continue
        subtopic = str(item.get("subtopic") or "").strip()
        if not subtopic:
            continue
        try:
            percent = int(item.get("percent", 0) or 0)
        except (TypeError, ValueError):
            percent = 0
        mastery[subtopic] = max(0, min(100, percent))

    order = practice_mastery_order(practice, fallback_order)
    for subtopic in mastery.keys():
        if subtopic not in order:
            order.append(subtopic)

    updated = {**practice, "mastery": mastery, "mastery_order": order}
    return updated


def practice_is_active(practice: dict[str, Any]) -> bool:
    return str(practice.get("status") or "").strip().casefold() in {
        "processing_summary",
        "summary_ready",
        "processing_quiz",
        "completed",
        "failed",
    }


def practice_state_from_patch(current_practice: dict[str, Any], patch: dict[str, Any]) -> dict[str, Any]:
    practice = normalize_practice_state(current_practice)
    practice.update(patch)
    practice["updated_at"] = now_iso()
    return practice


def invalidate_practice_state(reason: str) -> dict[str, Any]:
    state = default_practice_state()
    state["status"] = "stale"
    state["stale_reason"] = reason
    state["updated_at"] = now_iso()
    return state


def normalize_text_for_match(text: str) -> str:
    return re.sub(r"[^0-9a-zA-Zа-яА-ЯёЁ]+", " ", str(text or "").casefold()).strip()


def split_text_tokens(text: str) -> set[str]:
    tokens = {
        token
        for token in re.split(r"[^0-9a-zA-Zа-яА-ЯёЁ]+", normalize_text_for_match(text))
        if len(token) >= 4
    }
    return tokens


def select_practice_mini_summaries(generation: dict[str, Any], weak_subtopics: list[str]) -> list[dict[str, Any]]:
    mini_summaries = generation.get("mini_summary", [])
    summary_sections = generation.get("summary", [])
    if not isinstance(mini_summaries, list):
        mini_summaries = []
    if not isinstance(summary_sections, list):
        summary_sections = []

    summary_map: dict[str, dict[str, Any]] = {}
    for idx, section in enumerate(summary_sections):
        if not isinstance(section, dict):
            continue
        subtopic = str(section.get("subtopic") or f"Раздел {idx + 1}").strip()
        if subtopic:
            summary_map[normalize_text_for_match(subtopic)] = section

    selected: list[dict[str, Any]] = []
    seen_chunks: set[int] = set()

    def add_chunk(item: dict[str, Any]) -> None:
        try:
            chunk_id = int(item.get("chunk_id", 0) or 0)
        except (TypeError, ValueError):
            chunk_id = 0
        if chunk_id in seen_chunks:
            return
        seen_chunks.add(chunk_id)
        selected.append(item)

    for weak_subtopic in weak_subtopics:
        normalized_weak = normalize_text_for_match(weak_subtopic)
        if not normalized_weak:
            continue
        matched_section = summary_map.get(normalized_weak)
        matched_tokens = split_text_tokens(matched_section.get("content", "")) if isinstance(matched_section, dict) else set()
        if not matched_tokens:
            matched_tokens = split_text_tokens(weak_subtopic)

        for item in mini_summaries:
            if not isinstance(item, dict):
                continue
            blob_parts: list[str] = []
            for field in ("key_points", "terms", "examples"):
                value = item.get(field)
                if isinstance(value, list):
                    blob_parts.extend(str(part) for part in value if str(part).strip())
            blob = normalize_text_for_match(" ".join(blob_parts))
            if normalized_weak in blob or any(token in blob for token in matched_tokens):
                add_chunk(item)

    if not selected and isinstance(mini_summaries, list):
        for item in mini_summaries[:4]:
            if isinstance(item, dict):
                add_chunk(item)

    return selected


def build_practice_questions(generation: dict[str, Any], questions: list[dict[str, Any]], weak_subtopics: list[str]) -> list[dict[str, Any]]:
    quiz = generation.get("quiz", [])
    if not isinstance(quiz, list):
        quiz = []
    allowed = {normalize_text_for_match(item) for item in weak_subtopics if normalize_text_for_match(item)}
    normalized_questions: list[dict[str, Any]] = []

    quiz_by_id: dict[str, dict[str, Any]] = {}
    for idx, q in enumerate(quiz):
        if not isinstance(q, dict):
            continue
        qid = str(q.get("question_id", idx + 1))
        quiz_by_id[qid] = q

    for item in questions:
        if not isinstance(item, dict):
            continue
        subtopic = str(item.get("subtopic") or "").strip()
        if allowed and normalize_text_for_match(subtopic) not in allowed:
            continue
        qid = str(item.get("question_id") or "").strip()
        source_q = quiz_by_id.get(qid)
        source_question_type = str(source_q.get("question_type") if isinstance(source_q, dict) else "multiple_choice")
        source_question_text = str(source_q.get("question_text") if isinstance(source_q, dict) else "")
        source_explanation = str(source_q.get("explanation") if isinstance(source_q, dict) else "")
        normalized_questions.append(
            {
                "question_id": qid,
                "question_type": str(item.get("question_type") or source_question_type or "multiple_choice"),
                "subtopic": subtopic,
                "question_text": str(item.get("question_text") or source_question_text).strip(),
                "student_answer": str(item.get("student_answer") or "").strip(),
                "correct_answer": str(item.get("correct_answer") or "").strip(),
                "is_correct": bool(item.get("is_correct")),
                "explanation": str(item.get("explanation") or source_explanation).strip(),
            }
        )

    if not normalized_questions and questions:
        for item in questions:
            if not isinstance(item, dict):
                continue
            subtopic = str(item.get("subtopic") or "").strip()
            qid = str(item.get("question_id") or "").strip()
            source_q = quiz_by_id.get(qid)
            source_question_type = str(source_q.get("question_type") if isinstance(source_q, dict) else "multiple_choice")
            source_question_text = str(source_q.get("question_text") if isinstance(source_q, dict) else "")
            source_explanation = str(source_q.get("explanation") if isinstance(source_q, dict) else "")
            normalized_questions.append(
                {
                    "question_id": qid,
                    "question_type": str(item.get("question_type") or source_question_type or "multiple_choice"),
                    "subtopic": subtopic,
                    "question_text": str(item.get("question_text") or source_question_text).strip(),
                    "student_answer": str(item.get("student_answer") or "").strip(),
                    "correct_answer": str(item.get("correct_answer") or "").strip(),
                    "is_correct": bool(item.get("is_correct")),
                    "explanation": str(item.get("explanation") or source_explanation).strip(),
                }
            )

    return normalized_questions


def fallback_practice_questions_from_quiz(generation: dict[str, Any], weak_subtopics: list[str]) -> list[dict[str, Any]]:
    quiz = generation.get("quiz", [])
    if not isinstance(quiz, list):
        quiz = []
    allowed = {normalize_text_for_match(item) for item in weak_subtopics if normalize_text_for_match(item)}
    questions: list[dict[str, Any]] = []
    for idx, q in enumerate(quiz):
        if not isinstance(q, dict):
            continue
        subtopic = str(q.get("subtopic") or f"Подтема {idx + 1}").strip()
        normalized_subtopic = normalize_text_for_match(subtopic)
        if allowed and normalized_subtopic not in allowed:
            continue
        questions.append(
            {
                "question_id": str(q.get("question_id") or idx + 1),
                "question_type": str(q.get("question_type") or "multiple_choice"),
                "subtopic": subtopic,
                "question_text": str(q.get("question_text") or "").strip(),
                "student_answer": "",
                "correct_answer": str(q.get("correct_answer") or "").strip(),
                "is_correct": False,
                "explanation": str(q.get("explanation") or "").strip(),
            }
        )
    if not questions:
        for idx, q in enumerate(quiz):
            if not isinstance(q, dict):
                continue
            questions.append(
                {
                    "question_id": str(q.get("question_id") or idx + 1),
                    "question_type": str(q.get("question_type") or "multiple_choice"),
                    "subtopic": str(q.get("subtopic") or f"Подтема {idx + 1}").strip(),
                    "question_text": str(q.get("question_text") or "").strip(),
                    "student_answer": "",
                    "correct_answer": str(q.get("correct_answer") or "").strip(),
                    "is_correct": False,
                    "explanation": str(q.get("explanation") or "").strip(),
                }
            )
    return questions


def build_practice_payload(generation: dict[str, Any], weak_subtopics: list[str], questions: list[dict[str, Any]]) -> dict[str, Any]:
    topics = []
    selected_mini_summaries = select_practice_mini_summaries(generation, weak_subtopics)
    summary_sections = generation.get("summary", [])
    summary_lookup: dict[str, dict[str, Any]] = {}
    if isinstance(summary_sections, list):
        for idx, section in enumerate(summary_sections):
            if not isinstance(section, dict):
                continue
            subtopic = str(section.get("subtopic") or f"Раздел {idx + 1}").strip()
            if subtopic:
                summary_lookup[normalize_text_for_match(subtopic)] = section

    for weak_subtopic in weak_subtopics:
        normalized_weak = normalize_text_for_match(weak_subtopic)
        section = summary_lookup.get(normalized_weak)
        topic_mini_summaries = select_practice_mini_summaries(
            generation,
            [weak_subtopic],
        )
        topics.append(
            {
                "subtopic": weak_subtopic,
                "summary_section": section if isinstance(section, dict) else None,
                "mini_summaries": topic_mini_summaries,
            }
        )

    if not topics and selected_mini_summaries:
        topics.append(
            {
                "subtopic": "Повторение ключевых моментов",
                "summary_section": None,
                "mini_summaries": selected_mini_summaries,
            }
        )

    return {
        "weak_subtopics": weak_subtopics,
        "topics": topics,
        "questions": questions,
    }


async def grade_quiz_attempt(generation: dict[str, Any], quiz: list[dict[str, Any]], answers: list[dict[str, Any]]) -> dict[str, Any]:
    quiz_subtopics_list = quiz_subtopics(quiz)
    answers_by_id: dict[str, dict[str, Any]] = {}
    for item in answers:
        if not isinstance(item, dict):
            continue
        qid = str(item.get("question_id", "")).strip()
        if qid:
            answers_by_id[qid] = item

    results: list[dict[str, Any]] = []
    open_payload: list[dict[str, Any]] = []
    open_subtopics_by_id: dict[str, str] = {}
    for idx, q in enumerate(quiz):
        if not isinstance(q, dict):
            continue
        qid = str(q.get("question_id", idx + 1))
        qtype = q.get("question_type", "multiple_choice")
        subtopic = (q.get("subtopic") or f"Подтема {idx + 1}").strip()
        user_answer = answers_by_id.get(qid, {})
        if not isinstance(user_answer, dict):
            user_answer = {"answer": user_answer}
        if qtype in ("open_ended", "open_question"):
            open_subtopics_by_id[qid] = subtopic
            open_payload.append(
                {
                    "question_id": qid,
                    "question_text": q.get("question_text", ""),
                    "correct_answer": q.get("correct_answer", ""),
                    "student_answer": user_answer.get("student_answer") if isinstance(user_answer.get("student_answer"), str) else str(user_answer.get("answer") or ""),
                }
            )
            continue

        if "is_correct" in user_answer:
            score = 1 if user_answer.get("is_correct") is True else 0
        else:
            try:
                ua = int(user_answer.get("answer"))
            except (TypeError, ValueError):
                ua = -1
            try:
                correct_answer = int(q.get("correct_answer", -999))
            except (TypeError, ValueError):
                correct_answer = -999
            score = 1 if ua == correct_answer else 0
        results.append({"question_id": qid, "subtopic": subtopic, "score": score})

    if open_payload:
        if not ML_API_KEY:
            raise HTTPException(status_code=500, detail="Сервис проверки не настроен.")
        ml_client = MLServiceClient(api_key=ML_API_KEY, base_url=ML_URL)
        try:
            graded = await ml_client.grade_open_answers(open_payload)
            for row in graded.get("scores", []):
                qid = str(row.get("question_id", ""))
                results.append(
                    {
                        "question_id": qid,
                        "subtopic": open_subtopics_by_id.get(qid, ""),
                        "score": int(row.get("score", 0)),
                    }
                )
        except MLServiceError as exc:
            raise HTTPException(status_code=500, detail=exc.user_message)

    mastery = build_mastery_from_results(results, quiz_subtopics_list)
    recommendations = build_recommendations_from_mastery(mastery)
    recommendation = summarize_recommendations(recommendations)
    subtopic_to_revise = choose_subtopic_to_revise(recommendations)
    return {
        "results": results,
        "mastery": mastery,
        "recommendations": recommendations,
        "recommendation": recommendation,
        "subtopic_to_revise": subtopic_to_revise,
    }


def update_practice_state(generation_id: str, patch: dict[str, Any]) -> None:
    current = get_generation(generation_id)
    if not current:
        return
    practice = normalize_practice_state(current.get("practice", {}))
    practice.update(patch)
    practice["updated_at"] = now_iso()
    update_generation(generation_id, {"practice": practice})


def seed_practice_mastery(practice: dict[str, Any], payload_mastery: Any, generation: dict[str, Any]) -> dict[str, Any]:
    current = normalize_practice_state(practice)
    mastery = normalize_mastery_map(current.get("mastery", {}))
    payload_mastery_map = normalize_mastery_map(payload_mastery)
    if payload_mastery_map:
        for subtopic, percent in payload_mastery_map.items():
            mastery.setdefault(subtopic, percent)
    if not mastery:
        analytics = generation.get("analytics", {})
        if isinstance(analytics, dict):
            mastery = normalize_mastery_map(analytics.get("mastery", []))
    if mastery:
        order = practice_mastery_order(current, [str(item) for item in quiz_subtopics(generation.get("quiz", []))])
        for subtopic in mastery.keys():
            if subtopic not in order:
                order.append(subtopic)
        current["mastery"] = mastery
        current["mastery_order"] = order
    return current


def practice_completion_view(practice: dict[str, Any]) -> dict[str, Any]:
    return {
        "practice": normalize_practice_state(practice),
        "next_action": "done"
        if bool(practice.get("practice_completed"))
        else ("continue" if practice.get("round_submitted") and practice.get("pending_weak_subtopics") else "start"),
    }


def practice_round_context(practice: dict[str, Any], generation: dict[str, Any], payload_mastery: Any = None) -> tuple[dict[str, Any], list[str], list[str], bool]:
    current = seed_practice_mastery(practice, payload_mastery, generation)
    mastery = normalize_mastery_map(current.get("mastery", {}))
    order = practice_mastery_order(current, [str(item) for item in quiz_subtopics(generation.get("quiz", []))])
    low_topics = practice_low_topics(mastery, order)
    selected = [str(item.get("subtopic") or "").strip() for item in low_topics[:2] if str(item.get("subtopic") or "").strip()]
    pending = [str(item.get("subtopic") or "").strip() for item in low_topics[2:] if str(item.get("subtopic") or "").strip()]
    all_done = not selected
    if selected:
        current["practice_round"] = int(current.get("practice_round") or 0) + 1
        current["current_weak_subtopics"] = selected
        current["pending_weak_subtopics"] = pending
        current["weak_subtopics"] = selected
        current["round_submitted"] = False
        current["practice_completed"] = False
    else:
        current["current_weak_subtopics"] = []
        current["pending_weak_subtopics"] = []
        current["weak_subtopics"] = []
        current["round_submitted"] = True
        current["practice_completed"] = True
        current["status"] = "completed"
        current["stage"] = "quiz"
        current["summary"] = []
        current["quiz"] = []
    current["mastery"] = mastery
    current["mastery_order"] = order
    current["updated_at"] = now_iso()
    return current, selected, pending, all_done


def update_generation(generation_id: str, patch: dict[str, Any], broadcast_event_type: str = "generation_updated") -> None:
    current = get_generation(generation_id)
    if not current:
        return
    merged = {**current, **patch}
    if ("summary" in patch or "quiz" in patch) and "practice" not in patch:
        current_practice = normalize_practice_state(current.get("practice", {}))
        if practice_is_active(current_practice):
            merged["practice"] = invalidate_practice_state("Практика устарела после изменения конспекта или теста.")
    conn = db_conn()
    conn.execute(
        """
        UPDATE generations
        SET status = ?, progress_percent = ?, transcript_json = ?, mini_summary_json = ?, summary_json = ?, quiz_json = ?, practice_json = ?, analytics_json = ?, error_message = ?
        WHERE id = ?
        """,
        (
            merged["status"],
            float(merged.get("progress_percent", 0) or 0),
            json.dumps(merged.get("transcript", []), ensure_ascii=False),
            json.dumps(merged.get("mini_summary", []), ensure_ascii=False),
            json.dumps(merged.get("summary", []), ensure_ascii=False),
            json.dumps(merged.get("quiz", []), ensure_ascii=False),
            json.dumps(normalize_practice_state(merged.get("practice", {})), ensure_ascii=False),
            json.dumps(merged.get("analytics", {}), ensure_ascii=False),
            merged.get("error_message", ""),
            generation_id,
        ),
    )
    conn.commit()
    conn.close()
    user_id = merged.get("creator_id") or merged.get("user_id")
    if user_id:
        try:
            loop = asyncio.get_running_loop()
            payload: dict[str, Any] = {
                "type": broadcast_event_type,
                "generation_id": generation_id,
                "status": merged.get("status", ""),
            }
            if broadcast_event_type == "generation_analytics_updated":
                analytics = merged.get("analytics", {}) if isinstance(merged.get("analytics"), dict) else {}
                payload["studentsCompleted"] = analytics.get("studentsCompleted", 0)
            loop.create_task(broadcast_to_user(user_id, payload))
        except RuntimeError:
            pass


def update_generation_progress(generation_id: str, progress_percent: float, broadcast_event_type: str = "generation_updated") -> None:
    current = get_generation(generation_id)
    if not current:
        return
    current["progress_percent"] = float(progress_percent)
    conn = db_conn()
    conn.execute(
        "UPDATE generations SET progress_percent = ? WHERE id = ?",
        (float(progress_percent), generation_id),
    )
    conn.commit()
    conn.close()
    user_id = current.get("creator_id") or current.get("user_id")
    if user_id:
        try:
            loop = asyncio.get_running_loop()
            payload: dict[str, Any] = {
                "type": broadcast_event_type,
                "generation_id": generation_id,
                "status": current.get("status", ""),
            }
            loop.create_task(broadcast_to_user(user_id, payload))
        except RuntimeError:
            pass


def make_user_error_message(exc: Exception) -> str:
    if isinstance(exc, MediaConversionError):
        return "Не удалось подготовить аудио из файла. Проверьте формат файла или попробуйте другой файл."
    if isinstance(exc, MLServiceError):
        combined = f"{exc.user_message} {exc}".lower()
        if "429" in combined or "rate" in combined:
            return "Rate limit reached"
        return exc.user_message
    text = str(exc).lower()
    if "rate" in text or "429" in text:
        return "Rate limit reached"
    if "timeout" in text:
        return "Превышено время ожидания ответа сервиса. Попробуйте позже."
    return "Не удалось завершить генерацию. Попробуйте позже."


class MediaConversionError(Exception):
    pass


def media_suffix(file_name: str) -> str:
    suffix = Path(file_name or "").suffix.lower()
    if not suffix or len(suffix) > 12:
        return ".media"
    return suffix


def convert_to_wav_audio(file_bytes: bytes, file_name: str) -> tuple[bytes, str, str]:
    safe_name = sanitize_uploaded_filename(file_name)
    with tempfile.TemporaryDirectory(prefix="upload_audio_", dir=DATA_DIR) as tmp_dir:
        input_path = Path(tmp_dir) / f"input{media_suffix(safe_name)}"
        output_path = Path(tmp_dir) / "audio.wav"
        input_path.write_bytes(file_bytes)

        cmd = [
            "ffmpeg",
            "-y",
            "-hide_banner",
            "-loglevel",
            "error",
            "-i",
            str(input_path),
            "-vn",
            "-ac",
            "1",
            "-ar",
            str(AUDIO_SAMPLE_RATE),
            "-c:a",
            "pcm_s16le",
            str(output_path),
        ]
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=300, check=False)
        except (OSError, subprocess.TimeoutExpired) as exc:
            raise MediaConversionError(f"ffmpeg conversion failed: {exc}") from exc

        if result.returncode != 0 or not output_path.exists() or output_path.stat().st_size == 0:
            raise MediaConversionError(f"ffmpeg conversion failed: {result.stderr.strip()}")

        return output_path.read_bytes(), f"{Path(safe_name or 'media').stem or 'media'}.wav", "audio/wav"


def format_timestamp(seconds: float) -> str:
    total_seconds = max(0, int(round(seconds)))
    minutes, secs = divmod(total_seconds, 60)
    hours, minutes = divmod(minutes, 60)
    if hours:
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    return f"{minutes:02d}:{secs:02d}"


def media_duration_seconds(media_path: Path) -> float:
    cmd = [
        "ffprobe",
        "-v",
        "error",
        "-show_entries",
        "format=duration",
        "-of",
        "default=noprint_wrappers=1:nokey=1",
        str(media_path),
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60, check=False)
    except (OSError, subprocess.TimeoutExpired) as exc:
        raise MediaConversionError(f"ffprobe failed: {exc}") from exc

    if result.returncode != 0:
        raise MediaConversionError(f"ffprobe failed: {result.stderr.strip()}")

    try:
        duration = float(result.stdout.strip())
    except ValueError as exc:
        raise MediaConversionError(f"ffprobe returned invalid duration: {result.stdout.strip()}") from exc
    if duration <= 0:
        raise MediaConversionError("Audio duration is empty")
    return duration


def detect_silence_ranges(media_path: Path) -> list[tuple[float, float]]:
    cmd = [
        "ffmpeg",
        "-hide_banner",
        "-loglevel",
        "info",
        "-i",
        str(media_path),
        "-af",
        f"silencedetect=noise={AUDIO_SILENCE_NOISE_DB}:d={AUDIO_SILENCE_MIN_SECONDS}",
        "-f",
        "null",
        "-",
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=300, check=False)
    except (OSError, subprocess.TimeoutExpired) as exc:
        raise MediaConversionError(f"ffmpeg silence detection failed: {exc}") from exc

    if result.returncode not in (0, 1) and result.stderr.strip():
        raise MediaConversionError(f"ffmpeg silence detection failed: {result.stderr.strip()}")

    silence_ranges: list[tuple[float, float]] = []
    silence_start: Optional[float] = None
    for line in result.stderr.splitlines():
        if "silence_start:" in line:
            try:
                silence_start = float(line.rsplit("silence_start:", 1)[1].strip().split()[0])
            except (IndexError, ValueError):
                silence_start = None
        elif "silence_end:" in line and silence_start is not None:
            try:
                silence_end = float(line.rsplit("silence_end:", 1)[1].strip().split()[0])
            except (IndexError, ValueError):
                continue
            if silence_end > silence_start:
                silence_ranges.append((silence_start, silence_end))
            silence_start = None
    return silence_ranges


def choose_chunk_end(
    start_seconds: float,
    duration_seconds: float,
    silence_ranges: list[tuple[float, float]],
    *,
    target_seconds: int,
    min_seconds: int,
    max_seconds: int,
) -> float:
    min_end = min(start_seconds + min_seconds, duration_seconds)
    max_end = min(start_seconds + max_seconds, duration_seconds)
    if max_end <= min_end:
        return duration_seconds

    target_end = min(start_seconds + target_seconds, max_end)
    best_boundary: Optional[float] = None
    best_score: Optional[float] = None

    for silence_start, silence_end in silence_ranges:
        if silence_end < min_end or silence_start > max_end:
            continue
        boundary = max(min_end, min(target_end, silence_end))
        boundary = max(boundary, silence_start)
        boundary = min(boundary, silence_end, max_end)
        if boundary < min_end or boundary > max_end:
            continue
        score = abs(boundary - target_end)
        if best_score is None or score < best_score:
            best_score = score
            best_boundary = boundary

    if best_boundary is not None:
        return best_boundary
    return target_end


def split_audio_into_chunks(audio_bytes: bytes, audio_name: str) -> list[dict[str, Any]]:
    chunk_seconds = max(AUDIO_CHUNK_MIN_SECONDS, min(AUDIO_CHUNK_SECONDS, AUDIO_CHUNK_MAX_SECONDS))
    min_seconds = max(10, min(AUDIO_CHUNK_MIN_SECONDS, chunk_seconds))
    max_seconds = max(min_seconds, min(AUDIO_CHUNK_MAX_SECONDS, 10 * 60))
    target_seconds = max(min_seconds, min(AUDIO_CHUNK_SECONDS, max_seconds))
    overlap_seconds = max(0, min(AUDIO_CHUNK_OVERLAP_SECONDS, 5, chunk_seconds - 1))

    with tempfile.TemporaryDirectory(prefix="audio_chunks_", dir=DATA_DIR) as tmp_dir:
        tmp_path = Path(tmp_dir)
        source_path = tmp_path / "source.wav"
        source_path.write_bytes(audio_bytes)
        duration = media_duration_seconds(source_path)
        silence_ranges = detect_silence_ranges(source_path)

        chunks = []
        start = 0.0
        chunk_id = 1
        while start < duration:
            remaining = duration - start
            if remaining <= max_seconds:
                end = duration
            else:
                end = choose_chunk_end(
                    start,
                    duration,
                    silence_ranges,
                    target_seconds=target_seconds,
                    min_seconds=min_seconds,
                    max_seconds=max_seconds,
                )
                if end < start + min_seconds:
                    end = min(start + target_seconds, start + max_seconds, duration)
                end = min(max(end, start + min_seconds), start + max_seconds, duration)
            chunk_path = tmp_path / f"chunk_{chunk_id:03d}.wav"
            cmd = [
                "ffmpeg",
                "-y",
                "-hide_banner",
                "-loglevel",
                "error",
                "-ss",
                f"{start:.3f}",
                "-i",
                str(source_path),
                "-t",
                f"{end - start:.3f}",
                "-vn",
                "-ac",
                "1",
                "-ar",
                str(AUDIO_SAMPLE_RATE),
                "-c:a",
                "pcm_s16le",
                str(chunk_path),
            ]
            try:
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=300, check=False)
            except (OSError, subprocess.TimeoutExpired) as exc:
                raise MediaConversionError(f"ffmpeg chunking failed: {exc}") from exc

            if result.returncode != 0 or not chunk_path.exists() or chunk_path.stat().st_size == 0:
                raise MediaConversionError(f"ffmpeg chunking failed: {result.stderr.strip()}")

            chunks.append(
                {
                    "chunk_id": chunk_id,
                    "start_seconds": start,
                    "end_seconds": end,
                    "start_time": format_timestamp(start),
                    "end_time": format_timestamp(end),
                    "filename": f"{Path(audio_name).stem}_chunk_{chunk_id:03d}.wav",
                    "mime_type": "audio/wav",
                    "bytes": chunk_path.read_bytes(),
                }
            )

            if end >= duration:
                break
            start = max(0.0, end - overlap_seconds)
            chunk_id += 1

        return chunks


async def broadcast_to_user(user_id: str, payload: dict[str, Any]) -> None:
    clients = WS_CONNECTIONS.get(user_id, set()).copy()
    if not clients:
        return
    message = json.dumps(payload, ensure_ascii=False)
    for ws in clients:
        try:
            await ws.send_text(message)
        except Exception:
            WS_CONNECTIONS.get(user_id, set()).discard(ws)


def build_analytics(
    generation_id: str,
    quiz: list[dict[str, Any]],
    speech_analysis: Optional[dict[str, Any]] = None,
    speech_analysis_error: str = "",
) -> dict[str, Any]:
    analytics = {
        "studentLink": f"/material/{generation_id}/",
        "studentsCompleted": 0,
        "mastery": [],
        "recommendations": [],
    }
    if isinstance(speech_analysis, dict) and speech_analysis:
        analytics["speech_analysis"] = speech_analysis
    if str(speech_analysis_error or "").strip():
        analytics["speech_analysis_error"] = str(speech_analysis_error).strip()
    return analytics


def merge_speech_analysis_into_analytics(analytics: dict[str, Any], speech_analysis: Optional[dict[str, Any]]) -> dict[str, Any]:
    merged = dict(analytics or {})
    if isinstance(speech_analysis, dict) and speech_analysis:
        merged["speech_analysis"] = speech_analysis
    return merged


def speech_analysis_from_generation(generation: dict[str, Any]) -> dict[str, Any]:
    analytics = generation.get("analytics") if isinstance(generation.get("analytics"), dict) else {}
    speech_analysis = analytics.get("speech_analysis") if isinstance(analytics, dict) else {}
    return speech_analysis if isinstance(speech_analysis, dict) else {}


def transcript_line_by_start_ms(transcript: list[dict[str, Any]], start_ms: Any) -> Optional[dict[str, Any]]:
    try:
        target = int(start_ms)
    except (TypeError, ValueError):
        return None
    for line in transcript:
        if not isinstance(line, dict):
            continue
        try:
            line_start = int(line.get("start_ms", 0) or 0)
        except (TypeError, ValueError):
            continue
        if line_start == target:
            return line
    return None


def transcript_lines_by_ms_range(
    transcript: list[dict[str, Any]],
    start_ms: Any,
    end_ms: Any = None,
) -> list[dict[str, Any]]:
    try:
        start_value = int(start_ms)
    except (TypeError, ValueError):
        return []
    try:
        end_value = int(end_ms if end_ms is not None else start_value)
    except (TypeError, ValueError):
        end_value = start_value
    if end_value < start_value:
        start_value, end_value = end_value, start_value
    matches: list[dict[str, Any]] = []
    for line in transcript:
        if not isinstance(line, dict):
            continue
        try:
            line_start = int(line.get("start_ms", 0) or 0)
        except (TypeError, ValueError):
            continue
        if start_value <= line_start <= end_value:
            matches.append(line)
    return matches


def xlsx_column_name(index: int) -> str:
    name = ""
    index = max(1, index)
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def xlsx_escape_text(value: Any) -> str:
    return xml_escape(str(value if value is not None else ""), {'"': '&quot;', "'": '&apos;'})


def sanitize_xlsx_sheet_name(name: str, fallback: str = "Лист") -> str:
    cleaned = re.sub(r"[\[\]\*\/\\\?:]", " ", str(name or fallback))
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    cleaned = cleaned[:31]
    return cleaned or fallback


def normalize_speech_title(title: Any) -> str:
    value = str(title or "").strip()
    if value == "Преподаватель активно задаёт вопросы студентам":
        return "Вопросы преподавателя"
    if value == "Преподаватель реагирует на ответы студентов и развивает обсуждение":
        return "Ответы студентов"
    return value


def xlsx_cell_xml(value: Any, row_num: int, col_num: int, *, style_id: Optional[int] = None) -> str:
    cell_ref = f"{xlsx_column_name(col_num)}{row_num}"
    style_attr = f' s="{style_id}"' if style_id is not None else ""
    text = xlsx_escape_text(value)
    return f'<c r="{cell_ref}" t="inlineStr"{style_attr}><is><t xml:space="preserve">{text}</t></is></c>'


def build_xlsx_sheet_xml(sheet: dict[str, Any]) -> str:
    def normalize_cell(cell: Any) -> dict[str, Any]:
        if isinstance(cell, dict):
            span_raw = cell.get("span", 1)
            try:
                span = max(1, int(span_raw or 1))
            except (TypeError, ValueError):
                span = 1
            return {
                "value": cell.get("value", ""),
                "style": cell.get("style"),
                "span": span,
            }
        return {"value": cell, "style": None, "span": 1}

    def normalize_row(row: Any) -> dict[str, Any]:
        if isinstance(row, dict):
            cells = row.get("cells")
            if not isinstance(cells, list):
                cells = [cells] if cells is not None else []
            return {"cells": cells, "height": row.get("height")}
        if isinstance(row, list):
            return {"cells": row, "height": None}
        return {"cells": [row], "height": None}

    rows = sheet.get("rows") if isinstance(sheet.get("rows"), list) else []
    columns = sheet.get("cols") if isinstance(sheet.get("cols"), list) else []
    page_setup = sheet.get("page_setup") if isinstance(sheet.get("page_setup"), dict) else {}

    row_xml: list[str] = []
    merge_refs: list[str] = []
    max_cols = max(1, len(columns))

    for row_num, row in enumerate(rows, start=1):
        row_spec = normalize_row(row)
        cells = []
        col_num = 1
        for cell in row_spec["cells"]:
            cell_spec = normalize_cell(cell)
            span = int(cell_spec["span"] or 1)
            cells.append(xlsx_cell_xml(cell_spec["value"], row_num, col_num, style_id=cell_spec["style"]))
            if span > 1:
                merge_refs.append(f'{xlsx_column_name(col_num)}{row_num}:{xlsx_column_name(col_num + span - 1)}{row_num}')
            col_num += span
        max_cols = max(max_cols, col_num - 1)
        row_attr = f' r="{row_num}"'
        if row_spec["height"] is not None:
            row_attr += f' ht="{row_spec["height"]}" customHeight="1"'
        row_xml.append(f'<row{row_attr}>{"".join(cells)}</row>')

    dimension = f"A1:{xlsx_column_name(max_cols)}{max(1, len(rows))}"
    cols_xml = ''
    if columns:
        cols_xml = '<cols>' + ''.join(
            f'<col min="{index}" max="{index}" width="{float(width):.2f}" customWidth="1"/>'
            for index, width in enumerate(columns, start=1)
        ) + '</cols>'
    merge_xml = ''
    if merge_refs:
        merge_xml = f'<mergeCells count="{len(merge_refs)}">' + ''.join(f'<mergeCell ref="{ref}"/>' for ref in merge_refs) + '</mergeCells>'

    orientation = str(page_setup.get("orientation") or "landscape")
    paper_size = int(page_setup.get("paperSize") or 9)
    fit_width = int(page_setup.get("fitToWidth") or 1)
    fit_height = int(page_setup.get("fitToHeight") or 1)

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheetPr><pageSetUpPr fitToPage="1"/></sheetPr>'
        f'<dimension ref="{dimension}"/>'
        '<sheetViews><sheetView workbookViewId="0"><selection activeCell="A1" sqref="A1"/></sheetView></sheetViews>'
        '<sheetFormatPr defaultRowHeight="18"/>'
        f'{cols_xml}'
        '<sheetData>'
        + "".join(row_xml)
        + '</sheetData>'
        f'{merge_xml}'
        '<printOptions horizontalCentered="0" verticalCentered="0" headings="0" gridLines="0"/>'
        '<pageMargins left="0.35" right="0.35" top="0.45" bottom="0.45" header="0.2" footer="0.2"/>'
        f'<pageSetup orientation="{orientation}" paperSize="{paper_size}" fitToWidth="{fit_width}" fitToHeight="{fit_height}"/>'
        '</worksheet>'
    )


def build_xlsx_styles_xml() -> str:
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts count="6">'
        '<font><sz val="10"/><color rgb="FF1E293B"/><name val="Arial"/><family val="2"/></font>'
        '<font><b/><sz val="10"/><color rgb="FFFFFFFF"/><name val="Arial"/><family val="2"/></font>'
        '<font><b/><sz val="11"/><color rgb="FF0F172A"/><name val="Arial"/><family val="2"/></font>'
        '<font><i/><sz val="10"/><color rgb="FF64748B"/><name val="Arial"/><family val="2"/></font>'
        '<font><b/><sz val="14"/><color rgb="FF0F172A"/><name val="Arial"/><family val="2"/></font>'
        '<font><b/><sz val="12"/><color rgb="FF0F172A"/><name val="Arial"/><family val="2"/></font>'
        '</fonts>'
        '<fills count="4">'
        '<fill><patternFill patternType="none"/></fill>'
        '<fill><patternFill patternType="solid"><fgColor rgb="FF0F766E"/><bgColor indexed="64"/></patternFill></fill>'
        '<fill><patternFill patternType="solid"><fgColor rgb="FFEFF6FF"/><bgColor indexed="64"/></patternFill></fill>'
        '<fill><patternFill patternType="solid"><fgColor rgb="FFF8FAFC"/><bgColor indexed="64"/></patternFill></fill>'
        '</fills>'
        '<borders count="2">'
        '<border><left/><right/><top/><bottom/><diagonal/></border>'
        '<border>'
        '<left style="thin"><color rgb="FFD9E2EC"/></left>'
        '<right style="thin"><color rgb="FFD9E2EC"/></right>'
        '<top style="thin"><color rgb="FFD9E2EC"/></top>'
        '<bottom style="thin"><color rgb="FFD9E2EC"/></bottom>'
        '<diagonal/>'
        '</border>'
        '</borders>'
        '<cellStyleXfs count="1">'
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>'
        '</cellStyleXfs>'
        '<cellXfs count="6">'
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" applyAlignment="1"><alignment vertical="top" wrapText="1"/></xf>'
        '<xf numFmtId="0" fontId="1" fillId="1" borderId="1" xfId="0" applyAlignment="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf>'
        '<xf numFmtId="0" fontId="4" fillId="3" borderId="1" xfId="0" applyAlignment="1"><alignment vertical="center" wrapText="1"/></xf>'
        '<xf numFmtId="0" fontId="5" fillId="2" borderId="1" xfId="0" applyAlignment="1"><alignment vertical="center" wrapText="1"/></xf>'
        '<xf numFmtId="0" fontId="0" fillId="3" borderId="1" xfId="0" applyAlignment="1"><alignment vertical="top" wrapText="1"/></xf>'
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" applyAlignment="1"><alignment vertical="top" wrapText="1"/></xf>'
        '</cellXfs>'
        '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        '</styleSheet>'
    )


def build_xlsx_bytes(worksheets: list[dict[str, Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    unique_sheets: list[dict[str, Any]] = []
    used_names: set[str] = set()
    for index, sheet in enumerate(worksheets, start=1):
        base_name = sanitize_xlsx_sheet_name(str(sheet.get("name") or f"Лист {index}"))
        name = base_name
        suffix = 2
        while name in used_names:
            trimmed = base_name[: max(1, 31 - len(f" ({suffix})"))].rstrip()
            name = f"{trimmed} ({suffix})"[:31]
            suffix += 1
        used_names.add(name)
        unique_sheets.append({
            "name": name,
            "rows": sheet.get("rows") if isinstance(sheet.get("rows"), list) else [],
            "cols": sheet.get("cols") if isinstance(sheet.get("cols"), list) else [],
            "page_setup": sheet.get("page_setup") if isinstance(sheet.get("page_setup"), dict) else {},
            "freeze_panes": sheet.get("freeze_panes") if isinstance(sheet.get("freeze_panes"), str) else "",
        })

    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.creator = "FastClass"
    wb.properties.lastModifiedBy = "FastClass"

    thin = Side(style="thin", color="FFD9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill(fill_type="solid", fgColor="FFD9EAF7")
    title_fill = PatternFill(fill_type="solid", fgColor="FFC7DDF4")
    note_fill = PatternFill(fill_type="solid", fgColor="FFF8EFB6")
    sub_fill = PatternFill(fill_type="solid", fgColor="FFF1F5F9")
    section_blue_fill = PatternFill(fill_type="solid", fgColor="FFE8F1FF")
    section_green_fill = PatternFill(fill_type="solid", fgColor="FFE9F9EE")
    section_purple_fill = PatternFill(fill_type="solid", fgColor="FFF0E9FF")
    section_orange_fill = PatternFill(fill_type="solid", fgColor="FFFFEFE3")
    score_red_fill = PatternFill(fill_type="solid", fgColor="FFE76F61")
    score_orange_fill = PatternFill(fill_type="solid", fgColor="FFF2B28F")
    score_green_fill = PatternFill(fill_type="solid", fgColor="FF87C88D")
    score_gray_fill = PatternFill(fill_type="solid", fgColor="FFE5E7EB")
    header_text_fill = PatternFill(fill_type="solid", fgColor="FFFCE7D2")

    styles = {
        0: {
            "font": Font(name="Arial", size=10, color="FF1E293B"),
            "fill": PatternFill(fill_type=None),
            "alignment": Alignment(vertical="top", wrap_text=True),
        },
        1: {
            "font": Font(name="Arial", size=10, bold=True, color="FF0F172A"),
            "fill": header_fill,
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        2: {
            "font": Font(name="Arial", size=14, bold=True, color="FF0F172A"),
            "fill": title_fill,
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        3: {
            "font": Font(name="Arial", size=10, italic=True, color="FF64748B"),
            "fill": note_fill,
            "alignment": Alignment(vertical="top", wrap_text=True),
        },
        4: {
            "font": Font(name="Arial", size=10, bold=True, color="FF0F172A"),
            "fill": sub_fill,
            "alignment": Alignment(vertical="center", wrap_text=True),
        },
        5: {
            "font": Font(name="Arial", size=10, color="FF1E293B"),
            "fill": PatternFill(fill_type=None),
            "alignment": Alignment(vertical="top", wrap_text=True),
        },
        6: {
            "font": Font(name="Arial", size=10, bold=True, color="FF0F172A"),
            "fill": header_text_fill,
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        7: {
            "font": Font(name="Arial", size=9, italic=True, color="FF1F2937"),
            "fill": note_fill,
            "alignment": Alignment(vertical="top", wrap_text=True),
        },
        8: {
            "font": Font(name="Arial", size=10, bold=True, color="FF0F172A"),
            "fill": section_blue_fill,
            "alignment": Alignment(vertical="center", wrap_text=True),
        },
        9: {
            "font": Font(name="Arial", size=10, bold=True, color="FF0F172A"),
            "fill": section_green_fill,
            "alignment": Alignment(vertical="center", wrap_text=True),
        },
        10: {
            "font": Font(name="Arial", size=10, bold=True, color="FF0F172A"),
            "fill": section_purple_fill,
            "alignment": Alignment(vertical="center", wrap_text=True),
        },
        11: {
            "font": Font(name="Arial", size=11, bold=True, color="FFFFFFFF"),
            "fill": score_red_fill,
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        12: {
            "font": Font(name="Arial", size=11, bold=True, color="FF1F2937"),
            "fill": score_orange_fill,
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        13: {
            "font": Font(name="Arial", size=11, bold=True, color="FFFFFFFF"),
            "fill": score_green_fill,
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        },
        14: {
            "font": Font(name="Arial", size=10, bold=True, color="FF0F172A"),
            "fill": score_gray_fill,
            "alignment": Alignment(vertical="center", wrap_text=True),
        },
        15: {
            "font": Font(name="Arial", size=10, bold=True, color="FF0F172A"),
            "fill": PatternFill(fill_type=None),
            "alignment": Alignment(vertical="center", wrap_text=True),
        },
    }

    def apply_style(cell, style_id: Optional[int]) -> None:
        spec = styles.get(int(style_id) if style_id is not None else 0, styles[0])
        cell.font = spec["font"]
        cell.fill = spec["fill"]
        cell.alignment = spec["alignment"]
        cell.border = border

    for sheet in unique_sheets:
        ws = wb.create_sheet(title=sheet["name"])
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = 9
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_margins.left = 0.35
        ws.page_margins.right = 0.35
        ws.page_margins.top = 0.45
        ws.page_margins.bottom = 0.45
        ws.page_margins.header = 0.2
        ws.page_margins.footer = 0.2

        rows = sheet["rows"]
        max_cols = max(6, len(sheet.get("cols") or []))
        text_widths = [12.0] * max_cols

        for row_num, row in enumerate(rows, start=1):
            row_spec = row if isinstance(row, dict) else {"cells": row if isinstance(row, list) else [row], "height": None}
            cells = row_spec.get("cells")
            if not isinstance(cells, list):
                cells = [cells] if cells is not None else []
            if row_spec.get("height") is not None:
                try:
                    ws.row_dimensions[row_num].height = float(row_spec["height"])
                except (TypeError, ValueError):
                    pass

            col_num = 1
            for cell in cells:
                if isinstance(cell, dict):
                    value = cell.get("value", "")
                    span_raw = cell.get("span", 1)
                    try:
                        span = max(1, int(span_raw or 1))
                    except (TypeError, ValueError):
                        span = 1
                    style_id = cell.get("style")
                else:
                    value = cell
                    span = 1
                    style_id = None

                if col_num > max_cols:
                    max_cols = col_num
                    text_widths.extend([12.0] * (max_cols - len(text_widths)))

                top_left = ws.cell(row=row_num, column=col_num, value=value)
                apply_style(top_left, style_id)

                if span > 1:
                    end_col = col_num + span - 1
                    ws.merge_cells(start_row=row_num, start_column=col_num, end_row=row_num, end_column=end_col)
                    max_cols = max(max_cols, end_col)
                    if len(text_widths) < max_cols:
                        text_widths.extend([12.0] * (max_cols - len(text_widths)))
                value_text = str(value or "").replace("\n", " ")
                estimated = min(42.0, max(12.0, (len(value_text) * 0.85) + 2.0))
                if span == 1:
                    idx = col_num - 1
                    text_widths[idx] = max(text_widths[idx], estimated)
                else:
                    span_width = estimated / span
                    for offset in range(span):
                        idx = col_num - 1 + offset
                        if idx >= len(text_widths):
                            text_widths.extend([12.0] * (idx - len(text_widths) + 1))
                        text_widths[idx] = max(text_widths[idx], min(28.0, span_width))
                col_num += span

        for col_idx in range(1, max_cols + 1):
            width = text_widths[col_idx - 1] if col_idx - 1 < len(text_widths) else 12.0
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 10.0), 42.0)

        freeze_panes = sheet.get("freeze_panes") or "A5"
        ws.freeze_panes = freeze_panes

    return _write_workbook_to_bytes(wb)


def _write_workbook_to_bytes(workbook: Any) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def normalize_speech_analysis_export_view(speech: Any) -> dict[str, Any]:
    question_type_aliases = {
        "rhetorical": "rhetorical",
        "риторический": "rhetorical",
        "checking_understanding": "checking_understanding",
        "проверка понимания": "checking_understanding",
        "quiz": "quiz",
        "викторина": "quiz",
        "clarifying": "clarifying",
        "уточняющий": "clarifying",
        "open_ended": "open_ended",
        "open-ended": "open_ended",
        "открытый": "open_ended",
        "factual": "factual",
        "фактический": "factual",
        "other": "other",
        "другой": "other",
    }

    def normalize_question_type(value: Any) -> str:
        normalized = str(value or "").strip().casefold()
        return question_type_aliases.get(normalized, "other" if normalized else "")

    def normalize_fragment(fragment: Any) -> dict[str, Any]:
        if not isinstance(fragment, dict):
            text = str(fragment or "").strip()
            return {"start_ms": 0, "end_ms": 0, "text": text}
        try:
            start_value = int(fragment.get("start_ms", 0) or 0)
        except (TypeError, ValueError):
            start_value = 0
        try:
            end_value = int(fragment.get("end_ms", start_value) or start_value)
        except (TypeError, ValueError):
            end_value = start_value
        if end_value < start_value:
            end_value = start_value
        normalized: dict[str, Any] = {
            "start_ms": start_value,
            "end_ms": end_value,
            "text": str(fragment.get("text") or "").strip(),
        }
        fragment_type = str(fragment.get("type") or "").strip()
        if fragment_type:
            normalized["type"] = fragment_type
        question_type = normalize_question_type(fragment.get("question_type"))
        if question_type:
            normalized["question_type"] = question_type
        return normalized

    def merge_fragments(primary: list[dict[str, Any]], secondary: list[dict[str, Any]]) -> list[dict[str, Any]]:
        merged: list[dict[str, Any]] = []
        seen: set[tuple[int, int, str, str, str]] = set()
        for fragment in primary + secondary:
            if not isinstance(fragment, dict):
                continue
            key = (
                int(fragment.get("start_ms", 0) or 0),
                int(fragment.get("end_ms", 0) or 0),
                str(fragment.get("text") or "").strip(),
                str(fragment.get("type") or "").strip(),
                normalize_question_type(fragment.get("question_type")),
            )
            if not key[2] or key in seen:
                continue
            seen.add(key)
            normalized = dict(fragment)
            qtype = normalize_question_type(normalized.get("question_type"))
            if qtype:
                normalized["question_type"] = qtype
            elif "question_type" in normalized:
                normalized.pop("question_type", None)
            merged.append(normalized)
        return merged

    def collect_chunk_fragments(chunk_analyses: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
        fragments: list[dict[str, Any]] = []
        for chunk in chunk_analyses:
            if not isinstance(chunk, dict):
                continue
            items = chunk.get(key)
            if not isinstance(items, list):
                continue
            fragments.extend(normalize_fragment(item) for item in items if isinstance(item, dict) and str(item.get("text") or "").strip())
        return fragments

    def collect_chunk_events(chunk_analyses: list[dict[str, Any]]) -> list[dict[str, Any]]:
        items: list[dict[str, Any]] = []
        for chunk in chunk_analyses:
            if not isinstance(chunk, dict):
                continue
            events = chunk.get("lesson_events")
            if not isinstance(events, list):
                continue
            for event in events:
                if not isinstance(event, dict):
                    continue
                try:
                    start_value = int(event.get("start_ms", 0) or 0)
                except (TypeError, ValueError):
                    start_value = 0
                items.append(
                    {
                        "time_ms": start_value,
                        "time": format_timestamp(start_value / 1000),
                        "title": str(event.get("title") or "").strip() or "Событие урока",
                        "description": str(event.get("description") or "").strip(),
                    }
                )
        items.sort(key=lambda item: int(item.get("time_ms", 0) or 0))
        return items

    speech_data = speech if isinstance(speech, dict) else {}
    chunk_analyses = speech_data.get("chunk_analyses") if isinstance(speech_data.get("chunk_analyses"), list) else []

    lesson_format_raw = speech_data.get("lesson_format") if isinstance(speech_data.get("lesson_format"), dict) else {}
    audience_engagement_raw = speech_data.get("audience_engagement") if isinstance(speech_data.get("audience_engagement"), dict) else {}
    lesson_structure_raw = speech_data.get("lesson_structure") if isinstance(speech_data.get("lesson_structure"), dict) else {}
    material_explanation_raw = speech_data.get("material_explanation") if isinstance(speech_data.get("material_explanation"), dict) else {}
    teacher_recommendation_raw = speech_data.get("teacher_recommendation") if isinstance(speech_data.get("teacher_recommendation"), dict) else {}
    flags_raw = speech_data.get("flags") if isinstance(speech_data.get("flags"), dict) else {}

    questions_raw = audience_engagement_raw.get("questions_to_students") if isinstance(audience_engagement_raw.get("questions_to_students"), dict) else {}
    answers_raw = audience_engagement_raw.get("student_answers") if isinstance(audience_engagement_raw.get("student_answers"), dict) else {}
    timeline_raw = lesson_structure_raw.get("step_by_step_explanation") if isinstance(lesson_structure_raw.get("step_by_step_explanation"), dict) else {}
    goals_raw = lesson_structure_raw.get("goals_and_summary") if isinstance(lesson_structure_raw.get("goals_and_summary"), dict) else {}
    examples_raw = material_explanation_raw.get("examples_and_analogies") if isinstance(material_explanation_raw.get("examples_and_analogies"), dict) else {}

    derived_questions = collect_chunk_fragments(chunk_analyses, "teacher_questions")
    derived_answers = collect_chunk_fragments(chunk_analyses, "student_answers")
    derived_examples = collect_chunk_fragments(chunk_analyses, "examples_and_analogies")
    derived_timeline = collect_chunk_events(chunk_analyses)
    derived_profanity: list[dict[str, Any]] = []
    derived_familiarity: list[dict[str, Any]] = []
    for chunk in chunk_analyses:
        if not isinstance(chunk, dict):
            continue
        chunk_flags = chunk.get("flags") if isinstance(chunk.get("flags"), dict) else {}
        profanity_items = chunk_flags.get("profanity") if isinstance(chunk_flags.get("profanity"), list) else []
        familiarity_items = chunk_flags.get("overly_familiar_tone") if isinstance(chunk_flags.get("overly_familiar_tone"), list) else []
        derived_profanity.extend(normalize_fragment(item) for item in profanity_items if isinstance(item, dict) and str(item.get("text") or "").strip())
        derived_familiarity.extend(normalize_fragment(item) for item in familiarity_items if isinstance(item, dict) and str(item.get("text") or "").strip())

    derived_intro: dict[str, Any] | None = None
    derived_summary: dict[str, Any] | None = None
    for chunk in chunk_analyses:
        if not isinstance(chunk, dict):
            continue
        chunk_goals = chunk.get("goals_and_summary") if isinstance(chunk.get("goals_and_summary"), dict) else {}
        intro = chunk_goals.get("intro") if isinstance(chunk_goals.get("intro"), dict) else {}
        summary = chunk_goals.get("summary") if isinstance(chunk_goals.get("summary"), dict) else {}
        if derived_intro is None and intro:
            derived_intro = {
                "present": bool(intro.get("present")),
                "start_ms": intro.get("start_ms"),
                "comment": str(intro.get("comment") or "").strip(),
            }
        elif derived_intro is not None and not derived_intro.get("present") and bool(intro.get("present")):
            derived_intro = {
                "present": True,
                "start_ms": intro.get("start_ms"),
                "comment": str(intro.get("comment") or "").strip(),
            }
        if derived_summary is None and summary:
            derived_summary = {
                "present": bool(summary.get("present")),
                "start_ms": summary.get("start_ms"),
                "comment": str(summary.get("comment") or "").strip(),
            }
        elif derived_summary is not None and not derived_summary.get("present") and bool(summary.get("present")):
            derived_summary = {
                "present": True,
                "start_ms": summary.get("start_ms"),
                "comment": str(summary.get("comment") or "").strip(),
            }

    lesson_format = {
        "format": str(lesson_format_raw.get("format") or "").strip() or (
            "Агрегированный анализ речи преподавателя"
            if chunk_analyses
            else "Формат занятия не определен"
        ),
        "comment": str(lesson_format_raw.get("comment") or "").strip() or (
            f"Проанализировано чанков: {len(chunk_analyses)}"
            if chunk_analyses
            else "Агрегированный анализ речи преподавателя готов."
        ),
    }

    questions = {
        "title": normalize_speech_title(questions_raw.get("title") or "Вопросы преподавателя"),
        "comment": str(questions_raw.get("comment") or "").strip(),
        "fragments": merge_fragments(
            [normalize_fragment(fragment) for fragment in (questions_raw.get("fragments") if isinstance(questions_raw.get("fragments"), list) else [])],
            derived_questions,
        ),
    }
    answers = {
        "title": normalize_speech_title(answers_raw.get("title") or "Ответы студентов"),
        "comment": str(answers_raw.get("comment") or "").strip(),
        "fragments": merge_fragments(
            [normalize_fragment(fragment) for fragment in (answers_raw.get("fragments") if isinstance(answers_raw.get("fragments"), list) else [])],
            derived_answers,
        ),
    }

    timeline = {
        "title": str(timeline_raw.get("title") or "Таймлайн урока").strip(),
        "timeline": [],
    }
    if isinstance(timeline_raw.get("timeline"), list) and timeline_raw.get("timeline"):
        for item in timeline_raw.get("timeline"):
            if not isinstance(item, dict):
                continue
            try:
                start_value = int(item.get("start_ms", 0) or 0)
            except (TypeError, ValueError):
                start_value = 0
            timeline["timeline"].append(
                {
                    "start_ms": start_value,
                    "time": format_timestamp(start_value / 1000),
                    "title": str(item.get("title") or "Событие урока").strip(),
                    "comment": str(item.get("description") or "").strip(),
                }
            )
    else:
        timeline["timeline"] = [
            {"start_ms": item["time_ms"], "time": item["time"], "title": item["title"], "comment": item["description"]}
            for item in derived_timeline
        ]

    intro_raw = goals_raw.get("intro") if isinstance(goals_raw.get("intro"), dict) else {}
    summary_raw = goals_raw.get("summary") if isinstance(goals_raw.get("summary"), dict) else {}
    goals = {
        "title": str(goals_raw.get("title") or "Цели и итоги урока").strip(),
        "intro": {
            "present": bool(intro_raw.get("present")) if intro_raw else bool(derived_intro.get("present")) if isinstance(derived_intro, dict) else False,
            "start_ms": intro_raw.get("start_ms") if intro_raw else (derived_intro.get("start_ms") if isinstance(derived_intro, dict) else None),
            "comment": str(intro_raw.get("comment") or "").strip() if intro_raw else (str(derived_intro.get("comment") or "").strip() if isinstance(derived_intro, dict) else ""),
        },
        "summary": {
            "present": bool(summary_raw.get("present")) if summary_raw else bool(derived_summary.get("present")) if isinstance(derived_summary, dict) else False,
            "start_ms": summary_raw.get("start_ms") if summary_raw else (derived_summary.get("start_ms") if isinstance(derived_summary, dict) else None),
            "comment": str(summary_raw.get("comment") or "").strip() if summary_raw else (str(derived_summary.get("comment") or "").strip() if isinstance(derived_summary, dict) else ""),
        },
    }

    examples = {
        "title": str(examples_raw.get("title") or "Примеры, аналогии и сторителлинг").strip(),
        "fragments": merge_fragments(
            [normalize_fragment(fragment) for fragment in (examples_raw.get("fragments") if isinstance(examples_raw.get("fragments"), list) else [])],
            derived_examples,
        ),
    }

    recommendation = {
        "title": str(teacher_recommendation_raw.get("title") or "Рекомендация преподавателю").strip(),
        "comment": str(teacher_recommendation_raw.get("comment") or "").strip(),
    }

    flags: dict[str, Any] = {}
    for key, fallback_title, derived_fragments in [
        ("profanity", "Ненормативная лексика", derived_profanity),
        ("overly_familiar_tone", "Панибратство", derived_familiarity),
    ]:
        block = flags_raw.get(key) if isinstance(flags_raw.get(key), dict) else {}
        fragments = block.get("fragments") if isinstance(block.get("fragments"), list) else []
        flags[key] = {
            "title": str(block.get("title") or fallback_title).strip(),
            "present": bool(block.get("present")) or bool(derived_fragments) or bool(fragments),
            "fragments": merge_fragments(
                [normalize_fragment(fragment) for fragment in fragments],
                derived_fragments,
            ),
        }

    return {
        "lesson_format": lesson_format,
        "audience_engagement": {
            "questions_to_students": questions,
            "student_answers": answers,
        },
        "lesson_structure": {
            "step_by_step_explanation": timeline,
            "goals_and_summary": goals,
        },
        "material_explanation": {
            "examples_and_analogies": examples,
        },
        "teacher_recommendation": recommendation,
        "flags": flags,
        "chunk_analyses": chunk_analyses,
    }


def build_speech_analysis_export_worksheets(generation: dict[str, Any]) -> list[dict[str, Any]]:
    transcript = generation.get("transcript") if isinstance(generation.get("transcript"), list) else []
    speech = speech_analysis_from_generation(generation)
    if not speech:
        return []
    speech = normalize_speech_analysis_export_view(speech)

    question_type_labels = {
        "rhetorical": "риторический",
        "checking_understanding": "проверка понимания",
        "quiz": "викторина",
        "clarifying": "уточняющий",
        "open_ended": "открытый",
        "factual": "фактический",
        "other": "другой",
    }

    def transcript_fragment_row(section: str, title: str, fragment: Any) -> list[str]:
        fragment_text = str(fragment.get("text") or "").strip() if isinstance(fragment, dict) else str(fragment or "").strip()
        if isinstance(fragment, dict):
            fragment_type = str(fragment.get("type") or "").strip().casefold()
            if fragment_type in {"example", "analogy", "metaphor", "storytelling"}:
                type_labels = {"example": "пример", "analogy": "аналогия", "metaphor": "метафора", "storytelling": "сторителлинг"}
                fragment_text = f"{fragment_text} ({type_labels[fragment_type]})"
            qtype = str(fragment.get("question_type") or "").strip().casefold()
            if qtype in question_type_labels:
                fragment_text = f"{fragment_text} [{question_type_labels[qtype]}]"
        start_ms_value: Any = fragment.get("start_ms") if isinstance(fragment, dict) else None
        end_ms_value: Any = fragment.get("end_ms") if isinstance(fragment, dict) else None
        lines = transcript_lines_by_ms_range(transcript, start_ms_value, end_ms_value)
        found = "Да" if lines else "Нет"
        try:
            start_value = int(start_ms_value or 0) if start_ms_value is not None else 0
        except (TypeError, ValueError):
            start_value = 0
        try:
            end_value = int(end_ms_value if end_ms_value is not None else start_value)
        except (TypeError, ValueError):
            end_value = start_value
        if end_value < start_value:
            start_value, end_value = end_value, start_value
        timestamp = f"{format_timestamp(start_value / 1000)}–{format_timestamp(end_value / 1000)}" if end_value > start_value else format_timestamp(start_value / 1000)
        transcript_text = " / ".join(str(line.get("text") or "").strip() for line in lines if str(line.get("text") or "").strip())
        return [section, title, fragment_text, found, timestamp, transcript_text]

    lesson_format = speech.get("lesson_format") if isinstance(speech.get("lesson_format"), dict) else {}
    audience_engagement = speech.get("audience_engagement") if isinstance(speech.get("audience_engagement"), dict) else {}
    lesson_structure = speech.get("lesson_structure") if isinstance(speech.get("lesson_structure"), dict) else {}
    material_explanation = speech.get("material_explanation") if isinstance(speech.get("material_explanation"), dict) else {}
    flags = speech.get("flags") if isinstance(speech.get("flags"), dict) else {}
    chunk_analyses = speech.get("chunk_analyses") if isinstance(speech.get("chunk_analyses"), list) else []

    questions = audience_engagement.get("questions_to_students") if isinstance(audience_engagement.get("questions_to_students"), dict) else {}
    student_answers = audience_engagement.get("student_answers") if isinstance(audience_engagement.get("student_answers"), dict) else {}
    timeline = lesson_structure.get("step_by_step_explanation") if isinstance(lesson_structure.get("step_by_step_explanation"), dict) else {}
    goals = lesson_structure.get("goals_and_summary") if isinstance(lesson_structure.get("goals_and_summary"), dict) else {}
    examples = material_explanation.get("examples_and_analogies") if isinstance(material_explanation.get("examples_and_analogies"), dict) else {}
    teacher_recommendation = speech.get("teacher_recommendation") if isinstance(speech.get("teacher_recommendation"), dict) else {}

    questions_fragments = questions.get("fragments") if isinstance(questions.get("fragments"), list) else []
    answers_fragments = student_answers.get("fragments") if isinstance(student_answers.get("fragments"), list) else []
    examples_fragments = examples.get("fragments") if isinstance(examples.get("fragments"), list) else []
    timeline_items = timeline.get("timeline") if isinstance(timeline.get("timeline"), list) else []
    profanity_block = flags.get("profanity") if isinstance(flags.get("profanity"), dict) else {}
    familiarity_block = flags.get("overly_familiar_tone") if isinstance(flags.get("overly_familiar_tone"), dict) else {}
    profanity_fragments = profanity_block.get("fragments") if isinstance(profanity_block.get("fragments"), list) else []
    familiarity_fragments = familiarity_block.get("fragments") if isinstance(familiarity_block.get("fragments"), list) else []

    def fmt_score(_value: float) -> str:
        return "-"

    def score_style(value: float) -> int:
        if value >= 1.5:
            return 13
        if value >= 0.5:
            return 12
        return 11

    def factual_label(score: float) -> str:
        if score >= 1.5:
            return "было"
        if score >= 0.5:
            return "частично"
        return "не было"

    def count_types(fragments: list[dict[str, Any]], types: set[str]) -> int:
        total = 0
        for fragment in fragments:
            if not isinstance(fragment, dict):
                continue
            if str(fragment.get("question_type") or "").strip().casefold() in types:
                total += 1
        return total

    def section_rows(title: str, fill_style: int, items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        block: list[dict[str, Any]] = [{"cells": [{"value": title, "style": fill_style, "span": 6}], "height": 22}]
        block.append({"cells": [{"value": h, "style": 1} for h in ["Компетенция", "Поведенческие индикаторы", "Проявление", "Фактические действия (было/не было)", "Средний балл", "Дополнительные комментарии"]], "height": 20})
        scores: list[float] = []
        for item in items:
            score = float(item.get("score", 0) or 0)
            scores.append(score)
            block.append({
                "cells": [
                    {"value": item.get("competence", ""), "style": 14, "span": 1},
                    {"value": item.get("indicator", ""), "style": 5, "span": 1},
                    {"value": item.get("manifestation", ""), "style": 5, "span": 1},
                    {"value": factual_label(score), "style": 5, "span": 1},
                    {"value": fmt_score(score), "style": score_style(score), "span": 1},
                    {"value": item.get("comment", ""), "style": 5, "span": 1},
                ],
                "height": int(item.get("height", 24) or 24),
            })
        avg = sum(scores) / len(scores) if scores else 0
        note = str(items[-1].get("summary_note") or "") if items else ""
        block.append({
            "cells": [
                {"value": "Средний балл секции", "style": 14, "span": 4},
                {"value": fmt_score(avg), "style": score_style(avg), "span": 1},
                {"value": note or "Оценка секции по сигналам анализа речи.", "style": 5, "span": 1},
            ],
            "height": 24,
        })
        block.append({"cells": [], "height": 6})
        return block

    total_questions = len(questions_fragments)
    total_answers = len(answers_fragments)
    total_examples = len(examples_fragments)
    total_events = len(timeline_items)
    total_flags = len(profanity_fragments) + len(familiarity_fragments)
    has_timeline = total_events > 0
    open_question_count = count_types(questions_fragments, {"open_ended", "clarifying", "factual"})
    checking_question_count = count_types(questions_fragments, {"checking_understanding", "quiz"})
    rhetorical_question_count = count_types(questions_fragments, {"rhetorical"})
    has_intro = bool(goals.get("intro", {}).get("present")) if isinstance(goals.get("intro"), dict) else False
    has_summary = bool(goals.get("summary", {}).get("present")) if isinstance(goals.get("summary"), dict) else False
    has_profanity = bool(profanity_block.get("present"))
    has_familiarity = bool(familiarity_block.get("present"))

    created_at = generation.get("created_at")
    date_text = ""
    if created_at:
        try:
            date_text = datetime.fromisoformat(str(created_at).replace("Z", "+00:00")).strftime("%d.%m.%Y")
        except Exception:
            date_text = str(created_at)

    def first_text(items: list[dict[str, Any]], fallback: str) -> str:
        for item in items:
            if not isinstance(item, dict):
                continue
            text = str(item.get("text") or "").strip()
            if text:
                return text
        return fallback

    rows: list[dict[str, Any]] = [
        {"cells": [{"value": "Чек-лист качества преподавания на занятии", "style": 2, "span": 6}], "height": 24},
        {"cells": [{"value": "Сделайте копию шаблона и заполните показатели на основе анализа речи преподавателя. Оставьте пустые ячейки там, где сигналов недостаточно для объективной оценки.", "style": 7, "span": 6}], "height": 36},
        {"cells": [{"value": "Дата формирования", "style": 14, "span": 1}, {"value": date_text, "style": 5, "span": 2}, {"value": "Файл", "style": 14, "span": 1}, {"value": str(generation.get("file_name") or "Без названия").strip(), "style": 5, "span": 2}], "height": 20},
        {"cells": [{"value": "Формат занятия", "style": 14, "span": 1}, {"value": str(lesson_format.get("format") or "Формат занятия не определен").strip(), "style": 5, "span": 2}, {"value": "Рекомендация", "style": 14, "span": 1}, {"value": str(teacher_recommendation.get("title") or "Рекомендация преподавателю").strip(), "style": 5, "span": 2}], "height": 20},
        {"cells": [{"value": "Комментарий", "style": 14, "span": 1}, {"value": str(lesson_format.get("comment") or "Агрегированный анализ речи преподавателя готов.").strip(), "style": 5, "span": 5}], "height": 28},
        {"cells": [], "height": 6},
        {"cells": [{"value": "Легенда", "style": 4, "span": 6}], "height": 20},
        {"cells": [{"value": "Баллы", "style": 1}, {"value": "Смысл", "style": 1}, {"value": "Комментарий", "style": 1}, {"value": "", "style": 1}, {"value": "", "style": 1}, {"value": "", "style": 1}], "height": 20},
        {"cells": [{"value": "0", "style": 11, "span": 1}, {"value": "Отсутствие или несоответствие требованиям", "style": 5, "span": 2}, {"value": "Если показатель нельзя подтвердить по анализу или наблюдается нарушение.", "style": 7, "span": 3}], "height": 22},
        {"cells": [{"value": "1", "style": 12, "span": 1}, {"value": "Частичное соответствие", "style": 5, "span": 2}, {"value": "Сигнал есть, но он неполный, редкий или выражен неустойчиво.", "style": 7, "span": 3}], "height": 22},
        {"cells": [{"value": "2", "style": 13, "span": 1}, {"value": "Полное соответствие", "style": 5, "span": 2}, {"value": "Показатель проявляется уверенно и повторяется в анализе речи.", "style": 7, "span": 3}], "height": 22},
        {"cells": [], "height": 8},
        {"cells": [{"value": "Анализ по индикаторам", "style": 4, "span": 6}], "height": 20},
    ]

    rows.extend(section_rows("Организация понятного процесса обучения", 8, [
        {
            "competence": "Организация обучения",
            "indicator": "Приветствие и вход в урок",
            "manifestation": first_text(timeline_items, "В начале урока фиксируется вход в тему"),
            "score": 2 if has_intro or has_timeline else 1,
            "comment": "В начале есть структурирующий сигнал о входе в занятие." if has_intro or has_timeline else "Явного сигнала входа в урок не видно.",
        },
        {
            "competence": "Постановка задач",
            "indicator": "Формулирует цели и задачи занятия",
            "manifestation": str(goals.get("intro", {}).get("comment") or "Цели занятия проговариваются").strip() if isinstance(goals.get("intro"), dict) else "Цели занятия проговариваются",
            "score": 2 if has_intro else 0,
            "comment": "В начале урока есть явный сигнал о целях и рамке." if has_intro else "Явного сигнала о целях занятия не видно.",
        },
        {
            "competence": "Тайм-менеджмент",
            "indicator": "Укладывается в запланированную структуру",
            "manifestation": f"Этапов урока: {total_events}",
            "score": 2 if total_events >= 2 else 1 if total_events else 0,
            "comment": "В таймлайне просматриваются этапы занятия." if total_events else "Структура по времени выражена слабо.",
        },
        {
            "competence": "Педагогическая гибкость",
            "indicator": "Меняет подачу по ситуации",
            "manifestation": f"Примеры и аналогии: {total_examples}",
            "score": 2 if total_examples else 1 if total_questions else 0,
            "comment": "Примеры и аналогии помогают адаптировать объяснение." if total_examples else "Данных о гибкой перестройке подачи немного.",
        },
    ]))

    rows.extend(section_rows("Этика преподавания", 8, [
        {
            "competence": "Уважение к студентам и коллегам",
            "indicator": "Справедливо и объективно относится к участникам",
            "manifestation": "Ненормативная лексика и панибратство не подтверждены" if not has_profanity and not has_familiarity else "Есть сигналы, требующие внимания",
            "score": 2 if not has_profanity and not has_familiarity else 0,
            "comment": "Тон выглядит профессиональным и дистанция выдержана." if not has_profanity and not has_familiarity else "Есть речевые сигналы для доработки.",
        },
        {
            "competence": "Создание благоприятной атмосферы",
            "indicator": "Студенты не боятся задавать вопросы и комментировать",
            "manifestation": f"Вопросов преподавателя: {total_questions}, ответов студентов: {total_answers}",
            "score": 2 if total_questions and total_answers and not has_profanity else 1 if total_questions else 0,
            "comment": "Есть диалог и ответы студентов, атмосфера поддерживается вопросами." if total_questions else "Для уверенного вывода не хватает признаков диалога.",
        },
    ]))

    rows.extend(section_rows("Качество материала и владение им", 8, [
        {
            "competence": "Подготовка к занятию",
            "indicator": "Понимает образовательные результаты и структуру материала",
            "manifestation": str(lesson_format.get("format") or "Структурированный анализ речи").strip(),
            "score": 2 if lesson_format else 1,
            "comment": str(lesson_format.get("comment") or "Анализ построен по структуре урока.").strip(),
        },
        {
            "competence": "Доходчивость",
            "indicator": "Объясняет сложные моменты доступным способом",
            "manifestation": f"Примеры/аналогии: {total_examples}; ответы студентов: {total_answers}",
            "score": 2 if total_examples else 1 if total_answers else 0,
            "comment": "Объяснение поддерживается примерами и короткими пояснениями." if total_examples else "Пока мало прямых сигналов о доступности объяснения.",
        },
        {
            "competence": "Актуальность и широта",
            "indicator": "Выходит за рамки сухого пересказа",
            "manifestation": first_text(examples_fragments, "Примеры и аналогии как контекстуализация материала"),
            "score": 2 if total_examples >= 2 else 1 if total_examples else 0,
            "comment": "Материал подаётся через живые примеры и контекст." if total_examples else "Сигналов о широте контекста немного.",
        },
        {
            "competence": "Реакция на вопросы",
            "indicator": "Отвечает и развивает ответы аудитории",
            "manifestation": f"Вопросы: {total_questions}, ответы: {total_answers}",
            "score": 2 if total_answers else 1 if total_questions else 0,
            "comment": "В анализе есть взаимодействие на вопрос-ответ." if total_answers else "Недостаточно ответных реплик студентов.",
        },
    ]))

    rows.extend(section_rows("Управление динамикой занятия", 8, [
        {
            "competence": "Организация взаимодействия",
            "indicator": "Вовлекает студентов в процесс обучения",
            "manifestation": f"Всего вопросов: {total_questions}",
            "score": 2 if total_questions else 0,
            "comment": f"Есть {total_questions} вопросов к аудитории." if total_questions else "Вопросов к аудитории не найдено.",
        },
        {
            "competence": "Умение слушать",
            "indicator": "Серьёзно подходит к ответам студентов",
            "manifestation": f"Ответов студентов: {total_answers}",
            "score": 2 if total_answers else 0,
            "comment": "Есть ответы студентов, на которые можно опереться." if total_answers else "В данных нет ответов студентов.",
        },
        {
            "competence": "Мониторинг понимания",
            "indicator": "Регулярно проверяет понимание материала",
            "manifestation": f"Вопросов на проверку понимания: {checking_question_count}",
            "score": 2 if checking_question_count else 1 if total_questions else 0,
            "comment": "Вопросы на понимание присутствуют." if checking_question_count else "Проверка понимания выражена слабо.",
        },
        {
            "competence": "Обратная связь",
            "indicator": "Даёт конструктивную обратную связь",
            "manifestation": "Ответы и уточнения фиксируются в анализе" if total_answers else "Сигналов обратной связи мало",
            "score": 2 if total_answers else 0,
            "comment": "Взаимодействие с ответами студентов есть." if total_answers else "Недостаточно данных для оценки обратной связи.",
        },
    ]))

    rows.extend(section_rows("Структура речи и языка", 8, [
        {
            "competence": "Структурирование презентации",
            "indicator": "Чётко и логично выстраивает материал",
            "manifestation": f"Этапов урока: {total_events}",
            "score": 2 if total_events else 1,
            "comment": "Таймлайн показывает последовательность изложения." if total_events else "Структура изложения просматривается слабо.",
        },
        {
            "competence": "Грамотность речи",
            "indicator": "Избегает слов-паразитов и грубых выражений",
            "manifestation": "Ненормативная лексика не подтверждена" if not has_profanity else "Есть сигналы на проверку",
            "score": 2 if not has_profanity else 0,
            "comment": "Речь выглядит аккуратной и профессиональной." if not has_profanity else "Есть сигналы на доработку речевой культуры.",
        },
        {
            "competence": "Темп речи",
            "indicator": "Держит темп, при котором студенты успевают воспринимать информацию",
            "manifestation": f"Событий урока: {total_events}; вопросных сигналов: {total_questions}",
            "score": 2 if total_events >= 2 else 1,
            "comment": "Есть распределение по этапам и паузы для проверки понимания." if total_events >= 2 else "Темп трудно оценить по имеющимся сигналам.",
        },
        {
            "competence": "Понятные языковые конструкции",
            "indicator": "Мысли выражаются ёмко и ясно",
            "manifestation": str(lesson_format.get("comment") or "Наблюдается структурированная подача").strip(),
            "score": 2 if total_examples or total_questions else 1,
            "comment": "Формулировки выглядят собранными и понятными." if total_examples or total_questions else "Сигналов о языковой ясности немного.",
        },
    ]))

    rows.extend(section_rows("Приёмы вовлечения аудитории", 8, [
        {
            "competence": "Управление вниманием аудитории",
            "indicator": "Задаёт вопросы, чтобы вовлечь студентов",
            "manifestation": f"Всего вопросов: {total_questions}",
            "score": 2 if total_questions else 0,
            "comment": "Вовлечение строится через вопросы к аудитории." if total_questions else "Вопросные механики не просматриваются.",
        },
        {
            "competence": "Открытые вопросы",
            "indicator": "Использует открытые/уточняющие вопросы",
            "manifestation": f"Открытых и уточняющих вопросов: {open_question_count}",
            "score": 2 if open_question_count else 1 if total_questions else 0,
            "comment": "Вопросы на разворачивание мысли присутствуют." if open_question_count else "Открытые вопросы не доминируют.",
        },
        {
            "competence": "Примеры и аналогии",
            "indicator": "Использует сторителлинг, примеры и аналогии",
            "manifestation": f"Примеров/аналогий: {total_examples}",
            "score": 2 if total_examples else 0,
            "comment": "Примеры и аналогии помогают удерживать внимание." if total_examples else "Примеры и аналогии не отмечены.",
        },
        {
            "competence": "Разнообразие механик",
            "indicator": "Использует несколько способов вовлечения",
            "manifestation": f"Вопросы: {total_questions}, примеры: {total_examples}, риторические вопросы: {rhetorical_question_count}",
            "score": 2 if total_questions and total_examples else 1 if total_questions or total_examples else 0,
            "comment": "В анализе виден микс вопросов и примеров." if total_questions and total_examples else "Набор приёмов пока ограничен.",
        },
    ]))

    rows.extend(section_rows("Итог и рекомендация", 8, [
        {
            "competence": "Рекомендация преподавателю",
            "indicator": "Итоговый комментарий модели",
            "manifestation": str(teacher_recommendation.get("title") or "Рекомендация преподавателю").strip(),
            "score": 2 if str(teacher_recommendation.get("comment") or "").strip() else 1,
            "comment": str(teacher_recommendation.get("comment") or "Комментарий отсутствует.").strip(),
            "summary_note": "Это финальный вывод по всем сигналам анализа речи.",
        },
    ]))

    numeric_scores: list[float] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        cells = row.get("cells") if isinstance(row.get("cells"), list) else []
        for cell in cells:
            if isinstance(cell, dict) and int(cell.get("style", 0) or 0) in {11, 12, 13}:
                try:
                    numeric_scores.append(float(str(cell.get("value") or 0).replace(",", ".")))
                except Exception:
                    pass

    overall_score = sum(numeric_scores) / len(numeric_scores) if numeric_scores else 0
    rows.append({"cells": [{"value": "Итоговый средний балл", "style": 14, "span": 4}, {"value": "-", "style": score_style(overall_score), "span": 1}, {"value": "Сводная оценка по всем индикаторам анализа речи.", "style": 5, "span": 1}], "height": 28})
    rows.append({"cells": [], "height": 6})

    return [
        {
            "name": "Анализ речи",
            "rows": rows,
            "cols": [20, 28, 30, 16, 12, 40],
            "page_setup": {"orientation": "landscape", "paperSize": 9, "fitToWidth": 1, "fitToHeight": 1},
            "freeze_panes": "A12",
        }
    ]


def build_speech_analysis_export_worksheets_from_payload(payload: dict[str, Any]) -> list[dict[str, Any]]:
    transcript = payload.get("transcript") if isinstance(payload.get("transcript"), list) else []
    speech = payload.get("speech_analysis") if isinstance(payload.get("speech_analysis"), dict) else {}
    if not speech:
        return []
    generation = {
        "transcript": transcript,
        "analytics": {"speech_analysis": speech},
    }
    return build_speech_analysis_export_worksheets_precise(generation)


def build_speech_analysis_export_worksheets_precise(generation: dict[str, Any]) -> list[dict[str, Any]]:
    transcript = generation.get("transcript") if isinstance(generation.get("transcript"), list) else []
    speech = speech_analysis_from_generation(generation)
    if not speech:
        return []
    speech = normalize_speech_analysis_export_view(speech)

    lesson_format = speech.get("lesson_format") if isinstance(speech.get("lesson_format"), dict) else {}
    audience_engagement = speech.get("audience_engagement") if isinstance(speech.get("audience_engagement"), dict) else {}
    lesson_structure = speech.get("lesson_structure") if isinstance(speech.get("lesson_structure"), dict) else {}
    material_explanation = speech.get("material_explanation") if isinstance(speech.get("material_explanation"), dict) else {}
    flags = speech.get("flags") if isinstance(speech.get("flags"), dict) else {}
    teacher_recommendation = speech.get("teacher_recommendation") if isinstance(speech.get("teacher_recommendation"), dict) else {}

    questions = audience_engagement.get("questions_to_students") if isinstance(audience_engagement.get("questions_to_students"), dict) else {}
    student_answers = audience_engagement.get("student_answers") if isinstance(audience_engagement.get("student_answers"), dict) else {}
    timeline = lesson_structure.get("step_by_step_explanation") if isinstance(lesson_structure.get("step_by_step_explanation"), dict) else {}
    goals = lesson_structure.get("goals_and_summary") if isinstance(lesson_structure.get("goals_and_summary"), dict) else {}
    examples = material_explanation.get("examples_and_analogies") if isinstance(material_explanation.get("examples_and_analogies"), dict) else {}

    questions_fragments = questions.get("fragments") if isinstance(questions.get("fragments"), list) else []
    answers_fragments = student_answers.get("fragments") if isinstance(student_answers.get("fragments"), list) else []
    examples_fragments = examples.get("fragments") if isinstance(examples.get("fragments"), list) else []
    timeline_items = timeline.get("timeline") if isinstance(timeline.get("timeline"), list) else []
    profanity_block = flags.get("profanity") if isinstance(flags.get("profanity"), dict) else {}
    familiarity_block = flags.get("overly_familiar_tone") if isinstance(flags.get("overly_familiar_tone"), dict) else {}
    profanity_fragments = profanity_block.get("fragments") if isinstance(profanity_block.get("fragments"), list) else []
    familiarity_fragments = familiarity_block.get("fragments") if isinstance(familiarity_block.get("fragments"), list) else []

    question_type_labels = {
        "rhetorical": "риторический",
        "checking_understanding": "проверка понимания",
        "quiz": "викторина",
        "clarifying": "уточняющий",
        "open_ended": "открытый",
        "factual": "фактический",
        "other": "другой",
    }
    question_type_order = ["checking_understanding", "open_ended", "clarifying", "quiz", "factual", "rhetorical", "other"]

    def clean(value: Any) -> str:
        return str(value or "").strip()

    def percentage(part: int, total: int) -> str:
        if total <= 0:
            return "0%"
        return f"{round((part / total) * 100)}%"

    def timestamp_range(fragment: Any) -> str:
        if not isinstance(fragment, dict):
            return ""
        try:
            start_value = int(fragment.get("start_ms", 0) or 0)
        except (TypeError, ValueError):
            start_value = 0
        try:
            end_value = int(fragment.get("end_ms", start_value) or start_value)
        except (TypeError, ValueError):
            end_value = start_value
        if end_value < start_value:
            start_value, end_value = end_value, start_value
        return f"{format_timestamp(start_value / 1000)}–{format_timestamp(end_value / 1000)}" if end_value > start_value else format_timestamp(start_value / 1000)

    def transcript_text(fragment: Any) -> str:
        if not isinstance(fragment, dict):
            return ""
        lines = transcript_lines_by_ms_range(transcript, fragment.get("start_ms"), fragment.get("end_ms"))
        return " / ".join(clean(line.get("text")) for line in lines if clean(line.get("text")))

    def fragment_manifestation(fragment: Any) -> str:
        if not isinstance(fragment, dict):
            return clean(fragment)
        text = clean(fragment.get("text"))
        ftype = clean(fragment.get("type")).casefold()
        if ftype in {"example", "analogy", "metaphor", "storytelling"}:
            return f"{text} ({ {'example': 'пример', 'analogy': 'аналогия', 'metaphor': 'метафора', 'storytelling': 'сторителлинг'}[ftype] })" if text else ""
        qtype = clean(fragment.get("question_type")).casefold()
        if qtype in question_type_labels and text:
            return f"{text} [{question_type_labels[qtype]}]"
        return text

    def example_type_label(fragment: Any) -> str:
        if not isinstance(fragment, dict):
            return "Пример"
        ftype = clean(fragment.get("type")).casefold()
        labels = {
            "example": "Пример",
            "analogy": "Аналогия",
            "metaphor": "Метафора",
            "storytelling": "Сторителлинг",
        }
        return labels.get(ftype, "Пример")

    def fragment_comment(fragment: Any) -> str:
        if not isinstance(fragment, dict):
            return ""
        parts = [piece for piece in [timestamp_range(fragment), transcript_text(fragment)] if piece]
        return " | ".join(parts)

    def join_comments(fragments: list[dict[str, Any]], limit: int = 3) -> str:
        parts: list[str] = []
        seen: set[str] = set()
        for fragment in fragments[:limit]:
            comment = fragment_comment(fragment)
            if comment and comment not in seen:
                seen.add(comment)
                parts.append(comment)
        return "\n".join(parts)

    def point(_value: int) -> str:
        return "-"

    def question_points(count: int) -> int:
        return 2 if count > 0 else 0

    def atmosphere_points(answers_count: int, questions_count: int) -> int:
        if questions_count <= 0:
            return 0
        ratio = answers_count / questions_count
        if ratio > 0.4:
            return 2
        if ratio > 0.2:
            return 1
        return 0

    created_at = generation.get("created_at")
    date_text = ""
    if created_at:
        try:
            date_text = datetime.fromisoformat(str(created_at).replace("Z", "+00:00")).strftime("%d.%m.%Y")
        except Exception:
            date_text = clean(created_at)

    total_questions = len(questions_fragments)
    total_answers = len(answers_fragments)
    total_examples = len(examples_fragments)
    total_events = len(timeline_items)
    has_intro = bool(goals.get("intro", {}).get("present")) if isinstance(goals.get("intro"), dict) else False
    has_summary = bool(goals.get("summary", {}).get("present")) if isinstance(goals.get("summary"), dict) else False
    has_profanity = bool(profanity_block.get("present"))
    has_familiarity = bool(familiarity_block.get("present"))

    question_type_counts: dict[str, int] = {}
    question_type_fragments: dict[str, list[dict[str, Any]]] = {}
    seen_question_keys: set[str] = set()
    for fragment in questions_fragments:
        if not isinstance(fragment, dict):
            continue
        unique_key = f"{fragment.get('start_ms', '')}|{fragment.get('end_ms', '')}|{clean(fragment.get('text'))}"
        if unique_key in seen_question_keys:
            continue
        seen_question_keys.add(unique_key)
        qtype = clean(fragment.get("question_type")).casefold() or "other"
        question_type_counts[qtype] = question_type_counts.get(qtype, 0) + 1
        question_type_fragments.setdefault(qtype, []).append(fragment)

    rows: list[dict[str, Any]] = [
        {"cells": [{"value": "Чек-лист качества преподавания на занятии", "style": 2, "span": 5}], "height": 24},
        {"cells": [{"value": "Заполняйте только по фактам из анализа. Пустые ячейки лучше оставлять пустыми, если в данных нет подтверждения.", "style": 7, "span": 5}], "height": 34},
        {"cells": [{"value": "Дата формирования", "style": 14, "span": 1}, {"value": date_text, "style": 5, "span": 2}, {"value": "Файл", "style": 14, "span": 1}, {"value": clean(generation.get("file_name")), "style": 5, "span": 1}], "height": 20},
        {"cells": [{"value": "Формат занятия", "style": 14, "span": 1}, {"value": clean(lesson_format.get("format")), "style": 5, "span": 4}], "height": 20},
        {"cells": [], "height": 6},
        {"cells": [{"value": "Легенда", "style": 4, "span": 5}], "height": 20},
        {"cells": [{"value": "Баллы", "style": 1}, {"value": "Смысл", "style": 1}, {"value": "Проявление", "style": 1}, {"value": "", "style": 1}, {"value": "", "style": 1}], "height": 20},
        {"cells": [{"value": "0", "style": 11}, {"value": "Отсутствие или несоответствие", "style": 5}, {"value": "Показатель не подтвержден", "style": 5}, {"value": "", "style": 5}, {"value": "", "style": 5}], "height": 22},
        {"cells": [{"value": "1", "style": 12}, {"value": "Частичное соответствие", "style": 5}, {"value": "Показатель подтвержден частично", "style": 5}, {"value": "", "style": 5}, {"value": "", "style": 5}], "height": 22},
        {"cells": [{"value": "2", "style": 13}, {"value": "Полное соответствие", "style": 5}, {"value": "Показатель подтвержден", "style": 5}, {"value": "", "style": 5}, {"value": "", "style": 5}], "height": 22},
        {"cells": [], "height": 8},
        {"cells": [{"value": "Анализ по индикаторам", "style": 4, "span": 5}], "height": 20},
    ]

    def add_section(title: str, rows_to_add: list[dict[str, Any]]) -> None:
        if not rows_to_add:
            return
        rows.append({"cells": [{"value": title, "style": 8, "span": 5}], "height": 22})
        rows.append({"cells": [{"value": "Компетенция", "style": 1}, {"value": "Поведенческие индикаторы", "style": 1}, {"value": "Проявление", "style": 1}, {"value": "Баллы", "style": 1}, {"value": "Дополнительные комментарии", "style": 1}], "height": 20})
        rows.extend(rows_to_add)

    engagement_rows: list[dict[str, Any]] = []
    for qtype in question_type_order:
        count = question_type_counts.get(qtype, 0)
        if not count:
            continue
        fragments_for_type = question_type_fragments.get(qtype, [])
        engagement_rows.append({
            "cells": [
                {"value": "Вопросы преподавателя", "style": 14},
                {"value": question_type_labels.get(qtype, qtype), "style": 5},
                {"value": f"{count} из {total_questions} ({percentage(count, total_questions)})", "style": 5},
                {"value": point(question_points(count)), "style": 13 if count > 0 else 11},
                {"value": join_comments(fragments_for_type), "style": 5},
            ],
            "height": 26,
        })
    answers_comment = join_comments(answers_fragments) if answers_fragments else "Ответов студентов не было."
    atmosphere_score = atmosphere_points(total_answers, total_questions)
    engagement_rows.append({
        "cells": [
            {"value": "Ответы студентов", "style": 14},
            {"value": "Ответы студентов", "style": 5},
            {"value": f"{total_answers} из {total_questions} ({percentage(total_answers, total_questions)})", "style": 5},
            {"value": point(atmosphere_score), "style": 12 if atmosphere_score == 1 else 13 if atmosphere_score == 2 else 11},
            {"value": answers_comment, "style": 5},
        ],
        "height": 26,
    })
    add_section("Вовлечение аудитории", engagement_rows)

    structure_rows: list[dict[str, Any]] = []
    if total_events:
        timeline_comment = "\n".join(
            f"{clean(item.get('time') or '')} · {clean(item.get('title'))}{(' — ' + clean(item.get('comment'))) if clean(item.get('comment')) else ''}"
            for item in timeline_items
            if isinstance(item, dict)
        )
        structure_rows.append({
            "cells": [
                {"value": "Таймлайн урока", "style": 14},
                {"value": "Последовательность этапов", "style": 5},
                {"value": f"{total_events} событий", "style": 5},
                {"value": point(2 if total_events >= 2 else 1), "style": 13 if total_events >= 2 else 12},
                {"value": timeline_comment, "style": 5},
            ],
            "height": 34,
        })
    structure_rows.append({
        "cells": [
            {"value": "Цели и итоги урока", "style": 14},
            {"value": "Введение", "style": 5},
            {"value": "есть" if has_intro else "нет", "style": 5},
            {"value": point(2 if has_intro else 0), "style": 13 if has_intro else 11},
            {"value": clean(goals.get("intro", {}).get("comment") if isinstance(goals.get("intro"), dict) else ""), "style": 5},
        ],
        "height": 24,
    })
    structure_rows.append({
        "cells": [
            {"value": "Цели и итоги урока", "style": 14},
            {"value": "Завершение", "style": 5},
            {"value": "есть" if has_summary else "нет", "style": 5},
            {"value": point(2 if has_summary else 0), "style": 13 if has_summary else 11},
            {"value": clean(goals.get("summary", {}).get("comment") if isinstance(goals.get("summary"), dict) else ""), "style": 5},
        ],
        "height": 24,
    })
    add_section("Структура занятия", structure_rows)

    explanation_rows: list[dict[str, Any]] = []
    for fragment in examples_fragments:
        if not isinstance(fragment, dict):
            continue
        explanation_rows.append({
            "cells": [
                {"value": example_type_label(fragment), "style": 14},
                {"value": "Пример из речи преподавателя", "style": 5},
                {"value": fragment_manifestation(fragment), "style": 5},
                {"value": "-", "style": 5},
                {"value": fragment_comment(fragment), "style": 5},
            ],
            "height": 26,
        })
    add_section("Примеры, аналогии и сторителлинг", explanation_rows)

    flag_rows: list[dict[str, Any]] = []
    if has_profanity or profanity_fragments:
        flag_rows.append({
            "cells": [
                {"value": "Ненормативная лексика", "style": 14},
                {"value": "Тон и лексика", "style": 5},
                {"value": "не подтверждено" if not has_profanity else "есть сигналы", "style": 5},
                {"value": point(2 if not has_profanity and not profanity_fragments else 0), "style": 13 if not has_profanity and not profanity_fragments else 11},
                {"value": join_comments(profanity_fragments), "style": 5},
            ],
            "height": 26,
        })
    if has_familiarity or familiarity_fragments:
        flag_rows.append({
            "cells": [
                {"value": "Панибратство", "style": 14},
                {"value": "Обращение к аудитории", "style": 5},
                {"value": "не подтверждено" if not has_familiarity else "есть сигналы", "style": 5},
                {"value": point(2 if not has_familiarity and not familiarity_fragments else 0), "style": 13 if not has_familiarity and not familiarity_fragments else 11},
                {"value": join_comments(familiarity_fragments), "style": 5},
            ],
            "height": 26,
        })
    add_section("Флаги", flag_rows)

    recommendation_comment = clean(teacher_recommendation.get("comment"))
    recommendation_rows: list[dict[str, Any]] = []
    if recommendation_comment:
        recommendation_rows.append({
            "cells": [
                {"value": "Рекомендация преподавателю", "style": 14},
                {"value": "Комментарий", "style": 5},
                {"value": recommendation_comment, "style": 5},
                {"value": "", "style": 5},
                {"value": "", "style": 5},
            ],
            "height": 30,
        })
    add_section("Рекомендация преподавателю", recommendation_rows)

    return [
        {
            "name": "Анализ речи",
            "rows": rows,
            "cols": [22, 28, 34, 12, 44],
            "page_setup": {"orientation": "landscape", "paperSize": 9, "fitToWidth": 1, "fitToHeight": 1},
            "freeze_panes": "A12",
        }
    ]


def quiz_subtopics(quiz: list[dict[str, Any]]) -> list[str]:
    subtopics: list[str] = []
    for idx, q in enumerate(quiz):
        if not isinstance(q, dict):
            continue
        subtopic = str(q.get("subtopic") or f"Подтема {idx + 1}").strip()
        if subtopic and subtopic not in subtopics:
            subtopics.append(subtopic)
    return subtopics


def build_mastery_from_results(results: list[dict[str, Any]], subtopics: Optional[
    list[str]] = None) -> list[dict[str, Any]]:
    stats: dict[str, dict[str, int]] = {subtopic: {"correct": 0, "total": 0} for subtopic in (subtopics or [])}
    for item in results:
        subtopic = str(item.get("subtopic") or "Без темы").strip() or "Без темы"
        if subtopics and subtopic not in stats:
            continue
        current = stats.setdefault(subtopic, {"correct": 0, "total": 0})
        current["correct"] += 1 if item.get("score") else 0
        current["total"] += 1
    ordered_subtopics = subtopics or list(stats.keys())
    return [
        {
            "subtopic": subtopic,
            "percent": round((stat["correct"] / stat["total"]) * 100) if stat["total"] else 0,
            "correct": stat["correct"],
            "total": stat["total"],
        }
        for subtopic in ordered_subtopics
        for stat in [stats.get(subtopic, {"correct": 0, "total": 0})]
    ]


def analytics_from_attempts(generation_id: str, quiz: list[dict[str, Any]], attempts: list[sqlite3.Row]) -> dict[str, Any]:
    subtopics = quiz_subtopics(quiz)

    stats = {subtopic: {"correct": 0, "total": 0} for subtopic in subtopics}
    for attempt in attempts:
        try:
            results = json.loads(attempt["results_json"])
        except (TypeError, json.JSONDecodeError):
            results = []
        for item in results if isinstance(results, list) else []:
            subtopic = str(item.get("subtopic") or "Без темы").strip() or "Без темы"
            if subtopics and subtopic not in stats:
                continue
            current = stats.setdefault(subtopic, {"correct": 0, "total": 0})
            current["correct"] += 1 if item.get("score") else 0
            current["total"] += 1

    mastery = [
        {
            "subtopic": subtopic,
            "percent": round((stat["correct"] / stat["total"]) * 100) if stat["total"] else 0,
            "correct": stat["correct"],
            "total": stat["total"],
        }
        for subtopic in subtopics
        for stat in [stats.get(subtopic, {"correct": 0, "total": 0})]
        if stat["total"] > 0
    ]
    recommendations = build_recommendations_from_mastery(mastery)

    return {
        "studentLink": f"/material/{generation_id}/",
        "studentsCompleted": len(attempts),
        "mastery": mastery,
        "recommendations": recommendations,
    }


def build_recommendations_from_mastery(mastery: list[dict[str, Any]]) -> list[dict[str, Any]]:
    recommendations: list[dict[str, Any]] = []
    for item in mastery:
        try:
            percent = int(item.get("percent", 0) or 0)
        except (TypeError, ValueError):
            percent = 0
        subtopic = str(item.get("subtopic") or "Без темы").strip() or "Без темы"
        if percent < 50:
            recommendations.append(
                {
                    "subtopic": subtopic,
                    "action": "Важно разобрать тему",
                    "priority": "high",
                    "percent": percent,
                }
            )
        elif percent < 80:
            recommendations.append(
                {
                    "subtopic": subtopic,
                    "action": "Стоит повторить тему",
                    "priority": "medium",
                    "percent": percent,
                }
            )

    recommendations.sort(
        key=lambda item: (
            0 if item.get("priority") == "high" else 1,
            int(item.get("percent", 0) or 0),
            str(item.get("subtopic") or "").strip().casefold(),
        )
    )
    return recommendations[:2]


def summarize_recommendations(recommendations: list[dict[str, Any]]) -> str:
    if not recommendations:
        return "Отлично: все подтемы теста освоены."
    return " ".join(
        str(item.get("action") or "").strip()
        for item in recommendations
        if str(item.get("action") or "").strip()
    )


def choose_subtopic_to_revise(recommendations: list[dict[str, Any]]) -> str:
    for item in recommendations:
        if item.get("priority") == "high":
            return str(item.get("subtopic") or "").strip()
    return str(recommendations[0].get("subtopic") or "").strip() if recommendations else ""


def refresh_generation_analytics(generation_id: str) -> dict[str, Any]:
    generation = get_generation(generation_id)
    if not generation:
        return {}
    conn = db_conn()
    attempts = conn.execute(
        "SELECT * FROM student_attempts WHERE generation_id = ? ORDER BY created_at DESC",
        (generation_id,),
    ).fetchall()
    conn.close()
    analytics = analytics_from_attempts(generation_id, generation.get("quiz", []), attempts)
    if not attempts and not analytics.get("mastery"):
        analytics = build_analytics(generation_id, generation.get("quiz", []), speech_analysis_from_generation(generation))
        analytics["studentsCompleted"] = 0
        analytics["mastery"] = []
        analytics["recommendations"] = []
    else:
        analytics = merge_speech_analysis_into_analytics(analytics, speech_analysis_from_generation(generation))
    update_generation(generation_id, {"analytics": analytics}, broadcast_event_type="generation_analytics_updated")
    return analytics


def save_student_attempt(
    generation_id: str,
    user_id: str,
    answers: list[dict[str, Any]],
    results: list[dict[str, Any]],
    recommendation: str,
    subtopic_to_revise: str,
) -> dict[str, Any]:
    generation = get_generation(generation_id)
    mastery = build_mastery_from_results(results, quiz_subtopics(generation.get("quiz", [])) if generation else [])
    conn = db_conn()
    conn.execute(
        """
        INSERT INTO student_attempts
        (id, generation_id, user_id, created_at, answers_json, results_json, mastery_json, recommendation, subtopic_to_revise)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(generation_id, user_id) DO UPDATE SET
          created_at = excluded.created_at,
          answers_json = excluded.answers_json,
          results_json = excluded.results_json,
          mastery_json = excluded.mastery_json,
          recommendation = excluded.recommendation,
          subtopic_to_revise = excluded.subtopic_to_revise
        """,
        (
            f"attempt_{uuid.uuid4().hex[:14]}",
            generation_id,
            user_id,
            now_iso(),
            json.dumps(answers, ensure_ascii=False),
            json.dumps(results, ensure_ascii=False),
            json.dumps(mastery, ensure_ascii=False),
            recommendation,
            subtopic_to_revise,
        ),
    )
    conn.commit()
    conn.close()
    refresh_generation_analytics(generation_id)
    return {"mastery": mastery}


def load_student_attempt(generation_id: str, user_id: str) -> Optional[dict[str, Any]]:
    conn = db_conn()
    attempt_row = conn.execute(
        "SELECT * FROM student_attempts WHERE generation_id = ? AND user_id = ?",
        (generation_id, user_id),
    ).fetchone()
    conn.close()
    if not attempt_row:
        return None
    return {
        "answers": json.loads(attempt_row["answers_json"]) if attempt_row["answers_json"] else [],
        "results": json.loads(attempt_row["results_json"]) if attempt_row["results_json"] else [],
        "mastery": json.loads(attempt_row["mastery_json"]) if attempt_row["mastery_json"] else [],
        "recommendation": attempt_row["recommendation"] or "",
        "subtopic_to_revise": attempt_row["subtopic_to_revise"] or "",
        "created_at": attempt_row["created_at"],
    }


def shuffle_quiz_options(quiz: list[dict[str, Any]]) -> list[dict[str, Any]]:
    special_option_patterns = (
        "all of the above",
        "none of the above",
        "all of above",
        "все вышеперечисленное",
        "все выше перечисленное",
        "все перечисленное выше",
        "ни одно из вышеперечисленного",
        "ничего из вышеперечисленного",
        "ни один из вышеперечисленных",
        "ни один из перечисленных",
        "ни один из вышеперечисленных вариантов",
    )

    def is_special_option(option: Any) -> bool:
        text = str(option or "").strip().casefold()
        if not text:
            return False
        return any(pattern in text for pattern in special_option_patterns)

    shuffled_quiz = []
    for question in quiz:
        q = dict(question) if isinstance(question, dict) else {}
        options = q.get("options")
        if q.get("question_type") != "multiple_choice" or not isinstance(options, list) or len(options) < 2:
            shuffled_quiz.append(q)
            continue

        try:
            correct_idx = int(q.get("correct_answer"))
        except (TypeError, ValueError):
            shuffled_quiz.append(q)
            continue
        if correct_idx < 0 or correct_idx >= len(options):
            shuffled_quiz.append(q)
            continue

        normal_options = [(idx, option) for idx, option in enumerate(options) if not is_special_option(option)]
        special_options = [(idx, option) for idx, option in enumerate(options) if is_special_option(option)]

        random.shuffle(normal_options)
        ordered_options = normal_options + special_options
        q["options"] = [option for _, option in ordered_options]
        q["correct_answer"] = next(new_idx for new_idx, (old_idx, _) in enumerate(ordered_options) if old_idx == correct_idx)
        shuffled_quiz.append(q)
    return shuffled_quiz


def summary_subtopics(summary: list[dict[str, Any]]) -> list[str]:
    subtopics = []
    for idx, section in enumerate(summary):
        if not isinstance(section, dict):
            continue
        subtopic = str(section.get("subtopic") or f"Раздел {idx + 1}").strip()
        if subtopic and subtopic not in subtopics:
            subtopics.append(subtopic)
    return subtopics


def matching_summary_subtopic(candidate: str, subtopics: list[str]) -> str:
    normalized_candidate = candidate.strip().casefold()
    if not normalized_candidate:
        return ""
    for subtopic in subtopics:
        if subtopic.strip().casefold() == normalized_candidate:
            return subtopic
    return ""


def weakest_summary_subtopic(results: list[dict[str, Any]], subtopics: list[str]) -> str:
    stats: dict[str, dict[str, int]] = {}
    for item in results:
        subtopic = matching_summary_subtopic(str(item.get("subtopic") or ""), subtopics)
        if not subtopic:
            continue
        current = stats.setdefault(subtopic, {"correct": 0, "total": 0})
        current["correct"] += 1 if item.get("score") else 0
        current["total"] += 1
    weak = [
        (subtopic, stat["correct"] / stat["total"])
        for subtopic, stat in stats.items()
        if stat["total"] and stat["correct"] < stat["total"]
    ]
    return min(weak, key=lambda item: item[1])[0] if weak else ""


def transcript_to_chunks(transcript: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[int, list[dict[str, Any]]] = {}
    for phrase in transcript:
        try:
            chunk_id = int(phrase.get("chunk_id") or 1)
        except (TypeError, ValueError):
            chunk_id = 1
        grouped.setdefault(chunk_id, []).append(phrase)

    chunks = []
    for chunk_id, phrases in sorted(grouped.items()):
        phrases = sorted(phrases, key=lambda item: int(item.get("start_ms", 0) or 0))
        if not phrases:
            continue
        start_ms = int(phrases[0].get("start_ms", 0) or 0)
        end_ms = int(phrases[-1].get("start_ms", start_ms) or start_ms)
        chunks.append(
            {
                "chunk_id": chunk_id,
                "start_time": format_timestamp(start_ms / 1000),
                "end_time": format_timestamp(end_ms / 1000),
                "start_ms": start_ms,
                "end_ms": end_ms,
                "transcript": [
                    {
                        "start_ms": int(item.get("start_ms", 0) or 0),
                        "start_time": item.get("start_time") or format_timestamp(int(item.get("start_ms", 0) or 0) / 1000),
                        "text": item.get("text", ""),
                    }
                    for item in phrases
                ],
            }
        )
    return chunks


def _flatten_transcript_phrases(transcript_chunks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    phrases: list[dict[str, Any]] = []
    sorted_chunks = sorted(transcript_chunks, key=lambda item: int(item.get("start_ms", 0) or 0))
    for chunk in sorted_chunks:
        if not isinstance(chunk, dict):
            continue
        chunk_start_ms = int(chunk.get("start_ms", 0) or 0)
        transcript = chunk.get("transcript")
        if not isinstance(transcript, list):
            continue
        for phrase in transcript:
            if not isinstance(phrase, dict):
                continue
            text = str(phrase.get("text") or "").strip()
            if not text:
                continue
            try:
                start_ms = int(phrase.get("start_ms", chunk_start_ms) or chunk_start_ms)
            except (TypeError, ValueError):
                start_ms = chunk_start_ms
            phrases.append({"start_ms": start_ms, "text": text})
    phrases.sort(key=lambda item: int(item.get("start_ms", 0) or 0))
    return phrases


def _build_overlapped_time_groups(
    transcript_chunks: list[dict[str, Any]],
    *,
    target_seconds: int,
    min_seconds: int,
    max_seconds: int,
    overlap_phrases: int = 3,
) -> list[dict[str, Any]]:
    flat_phrases = _flatten_transcript_phrases(transcript_chunks)
    if not flat_phrases:
        return []

    target_ms = max(1, target_seconds) * 1000
    min_ms = max(1, min_seconds) * 1000
    max_ms = max(min_ms, max_seconds * 1000)
    overlap = max(0, int(overlap_phrases))

    def make_group(
        phrases: list[dict[str, Any]],
        chunk_id: int,
        *,
        start_ms: int | None = None,
        end_ms: int | None = None,
    ) -> dict[str, Any]:
        selected_start_ms = int(phrases[0].get("start_ms", 0) or 0)
        selected_end_ms = int(phrases[-1].get("start_ms", selected_start_ms) or selected_start_ms)
        start_value = selected_start_ms if start_ms is None else int(start_ms)
        end_value = selected_end_ms if end_ms is None else int(end_ms)
        if end_value < start_value:
            end_value = start_value
        return {
            "chunk_id": chunk_id,
            "start_time": format_timestamp(start_value / 1000),
            "end_time": format_timestamp(end_value / 1000),
            "start_ms": start_value,
            "end_ms": end_value,
            "transcript": [
                {
                    "start_ms": int(item.get("start_ms", start_value) or start_value),
                    "text": item.get("text", ""),
                }
                for item in phrases
            ],
        }

    def duration_ms(start_idx: int, end_idx: int) -> int:
        if end_idx <= start_idx:
            return 0
        start_ms = int(flat_phrases[start_idx].get("start_ms", 0) or 0)
        end_ms = int(flat_phrases[end_idx - 1].get("start_ms", start_ms) or start_ms)
        return max(0, end_ms - start_ms)

    def choose_cut_index(start_idx: int, end_idx: int, target_duration_ms: int) -> int:
        if end_idx - start_idx <= 1:
            return end_idx
        start_ms = int(flat_phrases[start_idx].get("start_ms", 0) or 0)
        end_ms = int(flat_phrases[end_idx - 1].get("start_ms", start_ms) or start_ms)
        span_ms = max(0, end_ms - start_ms)
        if span_ms <= 0:
            return start_idx + max(1, (end_idx - start_idx) // 2)
        target_cut_ms = start_ms + min(target_duration_ms, span_ms)
        best_idx = start_idx + 1
        best_distance = float("inf")
        for cut_idx in range(start_idx + 1, end_idx):
            cut_ms = int(flat_phrases[cut_idx].get("start_ms", start_ms) or start_ms)
            distance = abs(cut_ms - target_cut_ms)
            if distance < best_distance:
                best_distance = distance
                best_idx = cut_idx
        return best_idx

    core_ranges: list[tuple[int, int]] = []
    start_idx = 0
    while start_idx < len(flat_phrases):
        best_end = -1
        best_score = float("inf")
        end_idx = start_idx + 1

        while end_idx <= len(flat_phrases):
            current_duration_ms = duration_ms(start_idx, end_idx)
            if current_duration_ms > max_ms and end_idx > start_idx + 1:
                break
            if current_duration_ms >= min_ms:
                score = abs(current_duration_ms - target_ms)
                if score < best_score:
                    best_score = score
                    best_end = end_idx
            end_idx += 1

        if best_end < 0:
            remaining_duration_ms = duration_ms(start_idx, len(flat_phrases))
            if remaining_duration_ms <= max_ms and remaining_duration_ms > 0:
                best_end = len(flat_phrases)
            else:
                best_end = choose_cut_index(start_idx, len(flat_phrases), target_ms)

        if best_end <= start_idx:
            best_end = min(len(flat_phrases), start_idx + 1)

        core_ranges.append((start_idx, best_end))
        start_idx = best_end

    if len(core_ranges) > 1:
        last_start, last_end = core_ranges[-1]
        last_duration_ms = duration_ms(last_start, last_end)
        if last_duration_ms > 0 and last_duration_ms < min_ms:
            prev_start, prev_end = core_ranges[-2]
            combined_start = prev_start
            combined_end = last_end
            if duration_ms(combined_start, combined_end) <= max_ms:
                core_ranges[-2] = (combined_start, combined_end)
                core_ranges.pop()

    groups: list[dict[str, Any]] = []
    for idx, (core_start, core_end) in enumerate(core_ranges, start=1):
        payload_start = max(0, core_start - overlap)
        payload_end = min(len(flat_phrases), core_end + overlap)
        core_start_ms = int(flat_phrases[core_start].get("start_ms", 0) or 0)
        core_end_ms = int(flat_phrases[core_end - 1].get("start_ms", core_start_ms) or core_start_ms)
        groups.append(
            make_group(
                flat_phrases[payload_start:payload_end],
                idx,
                start_ms=core_start_ms,
                end_ms=core_end_ms,
            )
        )

    return groups


def transcript_to_summary_groups(
    transcript_chunks: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    return _build_overlapped_time_groups(
        transcript_chunks,
        target_seconds=SPEECH_ANALYSIS_GROUP_TARGET_SECONDS,
        min_seconds=SPEECH_ANALYSIS_GROUP_MIN_SECONDS,
        max_seconds=SPEECH_ANALYSIS_GROUP_MAX_SECONDS,
        overlap_phrases=SPEECH_ANALYSIS_GROUP_OVERLAP_PHRASES,
    )


def transcript_to_teacher_analysis_groups(transcript_chunks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return _build_overlapped_time_groups(
        transcript_chunks,
        target_seconds=SPEECH_ANALYSIS_GROUP_TARGET_SECONDS,
        min_seconds=SPEECH_ANALYSIS_GROUP_MIN_SECONDS,
        max_seconds=SPEECH_ANALYSIS_GROUP_MAX_SECONDS,
        overlap_phrases=SPEECH_ANALYSIS_GROUP_OVERLAP_PHRASES,
    )


def log_summary_payload(summary_groups: list[dict[str, Any]], label: str) -> None:
    print(f"[summary] {label}: groups={len(summary_groups)}")
    for idx, group in enumerate(summary_groups, start=1):
        subtopic = str(group.get("subtopic") or f"Раздел {idx}").strip()
        transcript = group.get("transcript", [])
        if isinstance(transcript, list):
            preview_parts = []
            for item in transcript[:2]:
                if not isinstance(item, dict):
                    continue
                text = " ".join(str(item.get("text", "") or "").split())
                if text:
                    preview_parts.append(text[:140])
            preview = " | ".join(preview_parts)
        else:
            preview = ""
        print(
            f"[summary] {label} #{idx}: subtopic={subtopic!r}, "
            f"start_ms={group.get('start_ms', 0)}, end_ms={group.get('end_ms', 0)}, preview={preview!r}"
        )


def log_final_summary(summary: list[dict[str, Any]], label: str) -> None:
    print(f"[summary] {label}: sections={len(summary) if isinstance(summary, list) else 0}")
    if not isinstance(summary, list):
        return
    for idx, section in enumerate(summary, start=1):
        if not isinstance(section, dict):
            continue
        subtopic = str(section.get("subtopic") or f"Раздел {idx}").strip()
        content = " ".join(str(section.get("content", "") or "").split())
        print(f"[summary] {label} #{idx}: subtopic={subtopic!r}, content_preview={content[:220]!r}")


def transcript_chunk_payloads(transcript_chunks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    payloads: list[dict[str, Any]] = []
    for chunk in transcript_chunks:
        if not isinstance(chunk, dict):
            continue
        if isinstance(chunk.get("transcript_chunk"), dict):
            payloads.append(chunk["transcript_chunk"])
        elif isinstance(chunk.get("transcript"), list):
            payloads.append(chunk)
    return payloads


def transcript_from_transcription_results(transcription_results: list[dict[str, Any]]) -> list[dict[str, Any]]:
    transcript: list[dict[str, Any]] = []
    for chunk in transcription_results:
        transcript.extend(chunk.get("phrases", []))
    transcript.sort(key=lambda item: (int(item.get("start_ms", 0) or 0), int(item.get("chunk_id", 0) or 0)))
    return transcript


def expand_transcript_segment(
    *,
    chunk_id: int,
    chunk_start_ms: int,
    chunk_end_ms: int,
    start_ms: int,
    text: str,
) -> list[dict[str, Any]]:
    clean_text = " ".join((text or "").split()).strip()
    if not clean_text:
        return []

    chunk_duration_ms = max(0, chunk_end_ms - chunk_start_ms)
    if chunk_duration_ms <= 0:
        return [
            {
                "chunk_id": chunk_id,
                "start_ms": max(chunk_start_ms, start_ms),
                "start_time": format_timestamp(max(chunk_start_ms, start_ms) / 1000),
                "text": clean_text,
                "is_final": True,
            }
        ]

    sentence_parts = [
        part.strip()
        for part in re.split(r"(?<=[.!?…])\s+(?=[A-ZА-ЯЁ0-9(«\"'])", clean_text)
        if part.strip()
    ]
    if len(sentence_parts) < 2:
        sentence_parts = [
            part.strip()
            for part in re.split(r"[\n\r]+", clean_text)
            if part.strip()
        ]
    if len(sentence_parts) < 2:
        return [
            {
                "chunk_id": chunk_id,
                "start_ms": max(chunk_start_ms, start_ms),
                "start_time": format_timestamp(max(chunk_start_ms, start_ms) / 1000),
                "text": clean_text,
                "is_final": True,
            }
        ]

    total_weight = sum(max(1, len(part)) for part in sentence_parts)
    expanded: list[dict[str, Any]] = []
    consumed_weight = 0
    for index, part in enumerate(sentence_parts):
        part_weight = max(1, len(part))
        part_start_ms = chunk_start_ms + round(chunk_duration_ms * consumed_weight / total_weight)
        if index > 0 and expanded:
            prev_start_ms = int(expanded[-1].get("start_ms", chunk_start_ms) or chunk_start_ms)
            if part_start_ms <= prev_start_ms:
                part_start_ms = prev_start_ms + 1
        if part_start_ms < chunk_start_ms:
            part_start_ms = chunk_start_ms
        expanded.append(
            {
                "chunk_id": chunk_id,
                "start_ms": part_start_ms,
                "start_time": format_timestamp(part_start_ms / 1000),
                "text": part,
                "is_final": True,
            }
        )
        consumed_weight += part_weight

    return expanded


async def transcribe_audio_chunks(
    ml_client: MLServiceClient,
    audio_chunks: list[dict[str, Any]],
    audio_content_type: str,
    generation_id: Optional[str] = None,
) -> list[dict[str, Any]]:
    async def animate_batch_progress(target_percent: int, stop_event: asyncio.Event) -> None:
        if not generation_id:
            return
        current_generation = get_generation(generation_id)
        current_percent = int(round(float(current_generation.get("progress_percent", 0) or 0))) if current_generation else 0
        current_percent = max(0, min(current_percent, target_percent))
        while current_percent < target_percent:
            if stop_event.is_set():
                return
            await asyncio.sleep(1)
            if stop_event.is_set():
                return
            current_percent += 1
            update_generation_progress(generation_id, current_percent)

    async def transcribe_one(chunk: dict[str, Any]) -> dict[str, Any]:
        chunk_start_ms = int(chunk["start_seconds"] * 1000)
        chunk_end_ms = int(chunk["end_seconds"] * 1000)
        chunk_duration_ms = max(0, chunk_end_ms - chunk_start_ms)
        chunk_segments = await ml_client.transcribe_chunk(
            file_name=chunk["filename"],
            mime_type=chunk.get("mime_type") or audio_content_type,
            audio_bytes=chunk["bytes"],
            chunk_id=chunk["chunk_id"],
            start_ms=chunk_start_ms,
            end_ms=chunk_end_ms,
        )
        chunk_phrases = []
        for segment in chunk_segments:
            phrase = str(segment.get("text", "")).strip()
            if not phrase:
                continue
            raw_start_ms = int(segment.get("start_ms", 0) or 0)
            if 0 <= raw_start_ms <= chunk_duration_ms + 5000:
                absolute_start_ms = chunk_start_ms + raw_start_ms
                if absolute_start_ms > chunk_end_ms + 2000:
                    absolute_start_ms = chunk_end_ms
            else:
                absolute_start_ms = raw_start_ms
            if absolute_start_ms < chunk_start_ms:
                absolute_start_ms = chunk_start_ms
            chunk_phrases.extend(
                expand_transcript_segment(
                    chunk_id=chunk["chunk_id"],
                    chunk_start_ms=chunk_start_ms,
                    chunk_end_ms=chunk_end_ms,
                    start_ms=absolute_start_ms,
                    text=phrase,
                )
            )
        transcript_chunk = {
            "chunk_id": chunk["chunk_id"],
            "start_time": chunk["start_time"],
            "end_time": chunk["end_time"],
            "start_ms": int(chunk["start_seconds"] * 1000),
            "end_ms": int(chunk["end_seconds"] * 1000),
            "transcript": [{"start_ms": phrase["start_ms"], "text": phrase["text"]} for phrase in chunk_phrases],
        }
        return {
            "chunk_id": chunk["chunk_id"],
            "transcript_chunk": transcript_chunk,
            "phrases": chunk_phrases,
        }

    results: list[dict[str, Any]] = []
    animated_progress = 0
    for batch_start in range(0, len(audio_chunks), TRANSCRIBE_BATCH_SIZE):
        batch = audio_chunks[batch_start:batch_start + TRANSCRIBE_BATCH_SIZE]
        batch_target = math.ceil(((batch_start + len(batch)) / len(audio_chunks)) * 100) if audio_chunks else 100
        batch_target = max(animated_progress, min(100, batch_target))
        stop_event = asyncio.Event()
        progress_task: Optional[asyncio.Task[None]] = None
        if generation_id:
            progress_task = asyncio.create_task(animate_batch_progress(batch_target, stop_event))
        try:
            batch_results = await asyncio.gather(*(transcribe_one(chunk) for chunk in batch))
        finally:
            if progress_task:
                stop_event.set()
                await progress_task
        results.extend(batch_results)
        if generation_id:
            animated_progress = batch_target
            update_generation(
                generation_id,
                {
                    "transcript": transcript_from_transcription_results(results),
                    "progress_percent": animated_progress,
                },
            )
    return sorted(results, key=lambda item: int(item.get("chunk_id", 0) or 0))


async def build_summary_and_quiz(
    ml_client: MLServiceClient,
    transcript_chunks: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    summary_source = transcript_chunk_payloads(transcript_chunks)
    summary_groups = transcript_to_summary_groups(summary_source)
    log_summary_payload(summary_groups, "build_summary_and_quiz")
    mini_summaries = await asyncio.gather(*(ml_client.make_mini_summary(chunk) for chunk in summary_groups))
    summary = await ml_client.make_lesson_summary(list(mini_summaries))
    log_final_summary(summary, "build_summary_and_quiz")
    quiz = shuffle_quiz_options(await ml_client.make_quiz(summary))
    transcript: list[dict[str, Any]] = []
    for chunk in transcript_chunks:
        if not isinstance(chunk, dict):
            continue
        phrases = chunk.get("phrases") if isinstance(chunk.get("phrases"), list) else chunk.get("transcript")
        if isinstance(phrases, list):
            transcript.extend(phrases)
    transcript.sort(key=lambda item: (int(item.get("start_ms", 0) or 0), int(item.get("chunk_id", 0) or 0)))
    return transcript, list(mini_summaries), summary, quiz


async def build_teacher_analysis(
    ml_client: MLServiceClient,
    transcript_chunks: list[dict[str, Any]],
) -> dict[str, Any]:
    analysis_source = transcript_to_teacher_analysis_groups(transcript_chunks)
    if not analysis_source:
        return {}
    log_summary_payload(analysis_source, "build_teacher_analysis")
    analysis_tasks = [asyncio.create_task(ml_client.make_teacher_analysis(chunk)) for chunk in analysis_source]
    chunk_analyses = await asyncio.gather(*analysis_tasks)
    aggregate = await ml_client.make_teacher_analysis_aggregate(list(chunk_analyses))
    if isinstance(aggregate, dict):
        aggregate["chunk_analyses"] = list(chunk_analyses)
    return aggregate if isinstance(aggregate, dict) else {}


async def build_summary_quiz_and_speech_analysis(
    ml_client: MLServiceClient,
    transcript_chunks: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any], str]:
    summary_task = asyncio.create_task(build_summary_and_quiz(ml_client, transcript_chunks))
    speech_task = asyncio.create_task(build_teacher_analysis(ml_client, transcript_chunks))
    try:
        transcript, mini_summaries, summary, quiz = await summary_task
    except Exception:
        speech_task.cancel()
        await asyncio.gather(speech_task, return_exceptions=True)
        raise

    speech_analysis: dict[str, Any] = {}
    speech_error = ""
    try:
        speech_analysis = await speech_task
    except Exception as exc:
        print("Teacher analysis failed:", exc)
        speech_analysis = {}
        speech_error = make_user_error_message(exc)

    return transcript, mini_summaries, summary, quiz, speech_analysis, speech_error


async def run_speech_analysis_retry_pipeline(generation_id: str) -> None:
    current = get_generation(generation_id)
    if not current:
        return
    if not ML_API_KEY:
        raise RuntimeError("ML_API_KEY is empty")
    transcript = current.get("transcript", [])
    transcript_chunks = transcript_to_chunks(transcript if isinstance(transcript, list) else [])
    analysis_source = transcript_to_teacher_analysis_groups(transcript_chunks)
    if not analysis_source:
        raise MLServiceError(
            "Retry requested without saved transcript",
            "Не найден сохраненный транскрипт для анализа речи преподавателя. Загрузите файл заново.",
        )

    ml_client = MLServiceClient(api_key=ML_API_KEY, base_url=ML_URL)
    existing_analytics = current.get("analytics") if isinstance(current.get("analytics"), dict) else {}
    if existing_analytics:
        analytics = dict(existing_analytics)
    else:
        analytics = build_analytics(generation_id, current.get("quiz", []))

    try:
        speech_analysis = await build_teacher_analysis(ml_client, transcript_chunks)
        if not speech_analysis:
            raise MLServiceError(
                "Teacher analysis retry returned empty result",
                "Не удалось заново собрать анализ речи преподавателя. Попробуйте еще раз.",
            )
        analytics["speech_analysis"] = speech_analysis
        analytics.pop("speech_analysis_error", None)
    except Exception as exc:
        analytics["speech_analysis_error"] = make_user_error_message(exc)

    update_generation(
        generation_id,
        {
            "status": "completed",
            "progress_percent": 100,
            "analytics": analytics,
            "error_message": "",
        },
    )


async def run_ml_retry_pipeline(generation_id: str) -> None:
    try:
        current = get_generation(generation_id)
        if not current:
            return
        if not ML_API_KEY:
            raise RuntimeError("ML_API_KEY is empty")

        ml_client = MLServiceClient(api_key=ML_API_KEY, base_url=ML_URL)
        summary = current.get("summary", [])
        quiz = current.get("quiz", [])
        transcript = current.get("transcript", [])
        transcript_chunks = transcript_to_chunks(transcript if isinstance(transcript, list) else [])
        summary_groups = transcript_to_summary_groups(transcript_chunks)
        if not summary_groups:
            raise MLServiceError(
                "Retry requested without saved transcript",
                "Не найден сохраненный транскрипт для повторной генерации. Загрузите файл заново.",
            )

        if isinstance(summary, list) and summary and isinstance(quiz, list) and quiz:
            await run_speech_analysis_retry_pipeline(generation_id)
            return

        update_generation(generation_id, {"status": "processing", "progress_percent": 100, "error_message": ""})
        log_summary_payload(summary_groups, "run_ml_retry_pipeline")
        transcript, mini_summaries, summary, quiz, speech_analysis, speech_error = await build_summary_quiz_and_speech_analysis(ml_client, transcript_chunks)
        log_final_summary(summary, "run_ml_retry_pipeline")
        analytics = build_analytics(generation_id, quiz, speech_analysis, speech_error)
        update_generation(
            generation_id,
            {
                "status": "completed",
                "progress_percent": 100,
                "mini_summary": list(mini_summaries),
                "summary": summary,
                "quiz": quiz,
                "transcript": transcript,
                "analytics": analytics,
                "error_message": "",
            },
        )
    except Exception as e:
        update_generation(generation_id, {"status": "failed", "error_message": make_user_error_message(e)})
        print("Generation retry failed:", e)


async def finalize_generation_from_transcript(generation_id: str, transcript: list[dict[str, Any]]) -> None:
    try:
        update_generation(
            generation_id,
            {
                "status": "processing",
                "progress_percent": 100,
                "mini_summary": [],
                "transcript": transcript,
                "summary": [],
                "quiz": [],
                "analytics": {},
                "error_message": "",
            },
        )
        if not ML_API_KEY:
            raise RuntimeError("ML_API_KEY is empty")
        ml_client = MLServiceClient(api_key=ML_API_KEY, base_url=ML_URL)
        transcript_chunks = transcript_to_chunks(transcript if isinstance(transcript, list) else [])
        summary_groups = transcript_to_summary_groups(transcript_chunks)
        if not summary_groups:
            raise MLServiceError("Cached transcript is empty", "Не удалось получить транскрипт из файла. Попробуйте другой файл.")

        log_summary_payload(summary_groups, "finalize_generation_from_transcript")
        transcript, mini_summaries, summary, quiz, speech_analysis, speech_error = await build_summary_quiz_and_speech_analysis(ml_client, transcript_chunks)
        log_final_summary(summary, "finalize_generation_from_transcript")
        analytics = build_analytics(generation_id, quiz, speech_analysis, speech_error)

        update_generation(
            generation_id,
            {
                "status": "completed",
                "progress_percent": 100,
                "transcript": transcript,
                "mini_summary": mini_summaries,
                "summary": summary,
                "quiz": quiz,
                "analytics": analytics,
                "error_message": "",
            },
        )
    except Exception as e:
        update_generation(generation_id, {"status": "failed", "error_message": make_user_error_message(e)})
        print("Generation from transcript failed:", e)


async def run_generation_pipeline(generation_id: str, file_bytes: bytes, file_name: str, content_type: Optional[str], content_hash: Optional[str] = None) -> None:
    try:
        update_generation(
            generation_id,
            {
                "status": "processing",
                "progress_percent": 0,
                "transcript": [],
                "mini_summary": [],
                "summary": [],
                "quiz": [],
                "analytics": {},
                "error_message": "",
            },
        )
        if not ML_API_KEY:
            raise RuntimeError("ML_API_KEY is empty")
        ml_client = MLServiceClient(api_key=ML_API_KEY, base_url=ML_URL)
        audio_bytes, audio_name, audio_content_type = await asyncio.to_thread(convert_to_wav_audio, file_bytes, file_name)
        audio_chunks = await asyncio.to_thread(split_audio_into_chunks, audio_bytes, audio_name)
        transcription_results = await transcribe_audio_chunks(ml_client, audio_chunks, audio_content_type, generation_id)
        if content_hash:
            store_cached_transcript(content_hash, transcript_from_transcription_results(transcription_results))

        transcript, mini_summaries, summary, quiz, speech_analysis, speech_error = await build_summary_quiz_and_speech_analysis(ml_client, transcription_results)
        analytics = build_analytics(generation_id, quiz, speech_analysis, speech_error)

        update_generation(
            generation_id,
            {
                "status": "completed",
                "progress_percent": 100,
                "transcript": transcript,
                "mini_summary": mini_summaries,
                "summary": summary,
                "quiz": quiz,
                "analytics": analytics,
                "error_message": "",
            },
        )
    except Exception as e:
        update_generation(generation_id, {"status": "failed", "error_message": make_user_error_message(e)})
        print("Generation failed:", e)


@app.get("/api/me")
async def api_me(request: Request):
    user_id = ensure_guest_user(request)
    return {"user_id": user_id}


@app.get("/api/generations")
async def api_generations(request: Request):
    user_id = ensure_guest_user(request)
    conn = db_conn()
    rows = conn.execute("SELECT * FROM generations WHERE COALESCE(creator_id, user_id) = ? ORDER BY created_at DESC LIMIT 20", (user_id,)).fetchall()
    conn.close()
    return {"items": [row_to_generation(r) for r in rows]}


@app.post("/api/generations/upload")
async def api_generations_upload(request: Request, background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    user_id = ensure_guest_user(request)
    try:
        file_bytes = await file.read()
        if not file_bytes:
            raise HTTPException(status_code=400, detail="Empty file")
        if len(file_bytes) > 200 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File too large")
        safe_file_name = sanitize_uploaded_filename(file.filename or "media")

        generation_id = f"gen_{uuid.uuid4().hex[:14]}"
        conn = db_conn()
        conn.execute(
            """
            INSERT INTO generations
            (id, user_id, creator_id, file_name, status, created_at, transcript_json, mini_summary_json, summary_json, quiz_json, practice_json, analytics_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (generation_id, user_id, user_id, safe_file_name, "processing", now_iso(), "[]", "[]", "[]", "[]", "{}", "{}"),
        )
        conn.commit()
        conn.close()

        content_hash = content_hash_for_bytes(file_bytes)
        cached_transcript = get_cached_transcript(content_hash)
        if cached_transcript:
            background_tasks.add_task(finalize_generation_from_transcript, generation_id, cached_transcript)
        else:
            background_tasks.add_task(run_generation_pipeline, generation_id, file_bytes, safe_file_name, file.content_type, content_hash)
        return JSONResponse(status_code=201, content={"id": generation_id, "content_hash": content_hash, "cache_hit": bool(cached_transcript)})
    finally:
        await file.close()


@app.post("/api/generations/{generation_id}/retry")
async def api_generation_retry(request: Request, background_tasks: BackgroundTasks, generation_id: str):
    user_id = ensure_guest_user(request)
    conn = db_conn()
    row = conn.execute("SELECT * FROM generations WHERE id = ? AND COALESCE(creator_id, user_id) = ?", (generation_id, user_id)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Not found")

    generation = row_to_generation(row)
    if generation.get("status") == "processing":
        raise HTTPException(status_code=409, detail="Генерация уже выполняется.")
    if not generation.get("transcript"):
        raise HTTPException(status_code=400, detail="Нет сохраненного транскрипта для повторной генерации.")

    await run_ml_retry_pipeline(generation_id)
    return {"ok": True}


@app.get("/api/generations/{generation_id}/speech-analysis.xlsx")
async def api_generation_speech_analysis_export(request: Request, generation_id: str):
    user_id = ensure_guest_user(request)
    conn = db_conn()
    row = conn.execute("SELECT * FROM generations WHERE id = ? AND COALESCE(creator_id, user_id) = ?", (generation_id, user_id)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Not found")

    generation = row_to_generation(row)
    worksheets = build_speech_analysis_export_worksheets_precise(generation)
    if not worksheets:
        raise HTTPException(status_code=404, detail="Speech analysis unavailable")

    xlsx_bytes = build_xlsx_bytes(worksheets)
    file_name = f"speech_analysis_{generation_id[:12]}.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


@app.post("/api/generations/{generation_id}/speech-analysis.xlsx")
async def api_generation_speech_analysis_export_from_payload(request: Request, generation_id: str):
    user_id = ensure_guest_user(request)
    conn = db_conn()
    row = conn.execute("SELECT * FROM generations WHERE id = ? AND COALESCE(creator_id, user_id) = ?", (generation_id, user_id)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Not found")

    try:
        payload = await request.json()
    except Exception:
        payload = {}
    if not isinstance(payload, dict):
        payload = {}

    worksheets = build_speech_analysis_export_worksheets_from_payload(payload)
    if not worksheets:
        generation = row_to_generation(row)
        worksheets = build_speech_analysis_export_worksheets_precise(generation)
    if not worksheets:
        raise HTTPException(status_code=404, detail="Speech analysis unavailable")

    xlsx_bytes = build_xlsx_bytes(worksheets)
    file_name = f"speech_analysis_{generation_id[:12]}.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
    )


@app.get("/api/generations/{generation_id}")
async def api_generation_get(request: Request, generation_id: str):
    user_id = ensure_guest_user(request)
    conn = db_conn()
    row = conn.execute("SELECT * FROM generations WHERE id = ? AND COALESCE(creator_id, user_id) = ?", (generation_id, user_id)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Not found")
    return row_to_generation(row)


@app.delete("/api/generations/{generation_id}")
async def api_generation_delete(request: Request, generation_id: str):
    user_id = ensure_guest_user(request)
    conn = db_conn()
    conn.execute("DELETE FROM generations WHERE id = ? AND COALESCE(creator_id, user_id) = ?", (generation_id, user_id))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.patch("/api/generations/{generation_id}")
async def api_generation_patch(request: Request, generation_id: str, payload: dict[str, Any]):
    user_id = ensure_guest_user(request)
    conn = db_conn()
    row = conn.execute("SELECT * FROM generations WHERE id = ? AND COALESCE(creator_id, user_id) = ?", (generation_id, user_id)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Not found")

    current = row_to_generation(row)
    if isinstance(payload.get("summary"), list):
        current["summary"] = payload["summary"]
    if isinstance(payload.get("quiz"), list):
        current["quiz"] = payload["quiz"]
    if isinstance(payload.get("analytics"), dict):
        current["analytics"] = payload["analytics"]
    update_generation(generation_id, current)
    return {"ok": True}


@app.get("/api/student/{generation_id}")
async def api_student(request: Request, generation_id: str):
    user_id = ensure_guest_user(request)
    generation = get_generation(generation_id)
    if not generation or generation.get("status") != "completed":
        raise HTTPException(status_code=404, detail="Not found")
    attempt = load_student_attempt(generation_id, user_id)
    return {
        "summary": generation.get("summary", []),
        "quiz": generation.get("quiz", []),
        "practice": generation.get("practice", default_practice_state()),
        "generation_id": generation_id,
        "attempt": attempt,
    }


@app.post("/api/student/{generation_id}/practice")
async def api_student_practice(request: Request, generation_id: str, payload: dict[str, Any]):
    ensure_guest_user(request)
    generation = get_generation(generation_id)
    if not generation or generation.get("status") != "completed":
        raise HTTPException(status_code=404, detail="Материал недоступен.")
    payload = payload if isinstance(payload, dict) else {}
    questions = payload.get("questions", [])
    if questions is not None and not isinstance(questions, list):
        raise HTTPException(status_code=400, detail="Некорректный формат вопросов.")

    current_practice = normalize_practice_state(generation.get("practice", {}))
    current_practice = seed_practice_mastery(current_practice, payload.get("mastery"), generation)
    if current_practice.get("practice_completed") and not current_practice.get("pending_weak_subtopics"):
        return practice_completion_view(current_practice)

    has_active_round = bool(current_practice.get("summary")) and not bool(current_practice.get("round_submitted"))
    if has_active_round and current_practice.get("status") in {"summary_ready", "processing_quiz", "completed"}:
        return practice_completion_view(current_practice)

    weak_subtopics = payload.get("weak_subtopics", [])
    if weak_subtopics is not None and not isinstance(weak_subtopics, list):
        raise HTTPException(status_code=400, detail="Нет слабых подтем для практики.")

    current_practice, selected_weak_subtopics, pending_weak_subtopics, all_done = practice_round_context(
        current_practice,
        generation,
        payload.get("mastery"),
    )
    practice_questions = build_practice_questions(
        generation,
        questions if isinstance(questions, list) else [],
        selected_weak_subtopics,
    )
    if not practice_questions:
        practice_questions = fallback_practice_questions_from_quiz(generation, selected_weak_subtopics)
        print(
            "[practice] fallback_questions",
            {
                "generation_id": generation_id,
                "questions_count": len(practice_questions),
                "used_selected_topics": bool(selected_weak_subtopics),
            },
        )

    print(
        "[practice] request",
        {
            "generation_id": generation_id,
            "weak_subtopics": current_practice.get("weak_subtopics", []),
            "selected_weak_subtopics": selected_weak_subtopics,
            "pending_weak_subtopics": pending_weak_subtopics,
            "questions_count": len(practice_questions),
            "all_done": all_done,
            "practice_round": int(current_practice.get("practice_round") or 0),
            "has_mastery": bool(current_practice.get("mastery")),
        },
    )

    if all_done:
        update_practice_state(
            generation_id,
            {
                "status": "completed",
                "stage": "quiz",
                "weak_subtopics": [],
                "current_weak_subtopics": [],
                "pending_weak_subtopics": [],
                "practice_round": int(current_practice.get("practice_round") or 0),
                "round_submitted": True,
                "practice_completed": True,
                "request": {
                    "weak_subtopics": [],
                    "questions": [],
                },
                "summary": [],
                "quiz": [],
                "error_message": "",
                "stale_reason": "",
                "mastery": current_practice.get("mastery", {}),
                "mastery_order": current_practice.get("mastery_order", []),
            },
        )
        practice = get_generation(generation_id)
        return practice_completion_view(practice.get("practice", default_practice_state()) if practice else default_practice_state())

    practice_request = {
        "weak_subtopics": selected_weak_subtopics,
        "questions": practice_questions,
        "mastery": current_practice.get("mastery", {}),
    }
    practice_context = build_practice_payload(generation, selected_weak_subtopics, practice_questions)
    print(
        "[practice] payload",
        {
            "weak_subtopics": practice_request["weak_subtopics"],
            "topics": practice_context.get("topics", []),
            "questions": practice_context.get("questions", []),
        },
    )
    update_practice_state(
        generation_id,
        {
            "status": "processing_summary",
            "stage": "summary",
            "weak_subtopics": selected_weak_subtopics,
            "current_weak_subtopics": selected_weak_subtopics,
            "pending_weak_subtopics": pending_weak_subtopics,
            "practice_round": int(current_practice.get("practice_round") or 0),
            "round_submitted": False,
            "practice_completed": False,
            "request": practice_request,
            "summary": [],
            "quiz": [],
            "error_message": "",
            "stale_reason": "",
            "mastery": current_practice.get("mastery", {}),
            "mastery_order": current_practice.get("mastery_order", []),
        },
    )

    if not ML_API_KEY:
        raise HTTPException(status_code=500, detail="Сервис дообучения не настроен.")

    ml_client = MLServiceClient(api_key=ML_API_KEY, base_url=ML_URL)
    try:
        practice_summary = await ml_client.make_practice_summary(practice_context)
        print(
            "[practice] ml response",
            {
                "generation_id": generation_id,
                "summary_count": len(practice_summary) if isinstance(practice_summary, list) else None,
                "summary_preview": practice_summary[:2] if isinstance(practice_summary, list) else practice_summary,
            },
        )
        update_practice_state(
            generation_id,
            {
                "status": "summary_ready",
                "stage": "summary",
                "weak_subtopics": selected_weak_subtopics,
                "current_weak_subtopics": selected_weak_subtopics,
                "pending_weak_subtopics": pending_weak_subtopics,
                "practice_round": int(current_practice.get("practice_round") or 0),
                "round_submitted": False,
                "practice_completed": False,
                "request": practice_request,
                "summary": practice_summary,
                "quiz": [],
                "error_message": "",
                "stale_reason": "",
                "mastery": current_practice.get("mastery", {}),
                "mastery_order": current_practice.get("mastery_order", []),
            },
        )
    except MLServiceError as exc:
        print(
            "[practice] ml error",
            {
                "generation_id": generation_id,
                "message": str(exc),
                "user_message": exc.user_message,
                "request": practice_context,
            },
        )
        update_practice_state(
            generation_id,
            {
                "status": "failed",
                "stage": "summary",
                "weak_subtopics": selected_weak_subtopics,
                "current_weak_subtopics": selected_weak_subtopics,
                "pending_weak_subtopics": pending_weak_subtopics,
                "practice_round": int(current_practice.get("practice_round") or 0),
                "round_submitted": False,
                "practice_completed": False,
                "request": practice_request,
                "summary": [],
                "quiz": [],
                "error_message": exc.user_message,
                "stale_reason": "",
                "mastery": current_practice.get("mastery", {}),
                "mastery_order": current_practice.get("mastery_order", []),
            },
        )
        raise HTTPException(status_code=500, detail=exc.user_message)

    practice = get_generation(generation_id)
    return practice_completion_view(practice.get("practice", default_practice_state()) if practice else default_practice_state())


@app.post("/api/student/{generation_id}/practice/quiz")
async def api_student_practice_quiz(request: Request, generation_id: str):
    ensure_guest_user(request)
    generation = get_generation(generation_id)
    if not generation or generation.get("status") != "completed":
        raise HTTPException(status_code=404, detail="Материал недоступен.")

    practice = normalize_practice_state(generation.get("practice", {}))
    if practice.get("status") == "completed" and isinstance(practice.get("quiz"), list) and practice.get("quiz"):
        return {"practice": practice}
    if practice.get("status") not in {"summary_ready", "failed", "processing_quiz"} or not isinstance(practice.get("summary"), list) or not practice.get("summary"):
        raise HTTPException(status_code=400, detail="Сначала нужно сгенерировать практический конспект.")

    if not ML_API_KEY:
        raise HTTPException(status_code=500, detail="Сервис дообучения не настроен.")

    update_practice_state(
        generation_id,
        {
            "status": "processing_quiz",
            "stage": "quiz",
            "error_message": "",
            "stale_reason": "",
        },
    )
    ml_client = MLServiceClient(api_key=ML_API_KEY, base_url=ML_URL)
    try:
        practice_quiz = shuffle_quiz_options(await ml_client.make_quiz(practice["summary"]))
        update_practice_state(
            generation_id,
            {
                "status": "completed",
                "stage": "quiz",
                "summary": practice["summary"],
                "quiz": practice_quiz,
                "weak_subtopics": practice.get("weak_subtopics", []),
                "request": practice.get("request", {}),
                "error_message": "",
                "stale_reason": "",
            },
        )
    except MLServiceError as exc:
        update_practice_state(
            generation_id,
            {
                "status": "failed",
                "stage": "quiz",
                "summary": practice["summary"],
                "quiz": [],
                "weak_subtopics": practice.get("weak_subtopics", []),
                "request": practice.get("request", {}),
                "error_message": exc.user_message,
                "stale_reason": "",
            },
        )
        raise HTTPException(status_code=500, detail=exc.user_message)

    practice = get_generation(generation_id)
    return {"practice": practice.get("practice", default_practice_state()) if practice else default_practice_state()}


@app.post("/api/student/{generation_id}/practice/complete")
async def api_student_practice_complete(request: Request, generation_id: str, payload: dict[str, Any]):
    ensure_guest_user(request)
    generation = get_generation(generation_id)
    if not generation or generation.get("status") != "completed":
        raise HTTPException(status_code=404, detail="Материал недоступен.")

    practice = normalize_practice_state(generation.get("practice", {}))
    if not isinstance(practice.get("quiz"), list) or not practice.get("quiz"):
        raise HTTPException(status_code=400, detail="Сначала нужно сгенерировать практический тест.")

    answers = payload.get("answers", []) if isinstance(payload, dict) else []
    if not isinstance(answers, list):
        raise HTTPException(status_code=400, detail="Некорректный формат ответов.")

    graded = await grade_quiz_attempt(generation, practice.get("quiz", []), answers)
    current_mastery = normalize_mastery_map(practice.get("mastery", {}))
    current_mastery = {**current_mastery}
    round_subtopics = quiz_subtopics(practice.get("quiz", []))
    round_mastery = build_mastery_from_results(graded["results"], round_subtopics)
    for item in round_mastery:
        subtopic = str(item.get("subtopic") or "").strip()
        if not subtopic:
            continue
        try:
            percent = int(item.get("percent", 0) or 0)
        except (TypeError, ValueError):
            percent = 0
        current_mastery[subtopic] = max(0, min(100, percent))

    order = practice_mastery_order(practice, round_subtopics)
    low_topics = practice_low_topics(current_mastery, order)
    pending_weak_subtopics = [str(item.get("subtopic") or "").strip() for item in low_topics if str(item.get("subtopic") or "").strip()]
    practice_completed = not pending_weak_subtopics
    practice_patch = {
        "status": "completed",
        "stage": "quiz",
        "weak_subtopics": practice.get("current_weak_subtopics", practice.get("weak_subtopics", [])),
        "current_weak_subtopics": practice.get("current_weak_subtopics", practice.get("weak_subtopics", [])),
        "pending_weak_subtopics": pending_weak_subtopics,
        "practice_round": int(practice.get("practice_round") or 0),
        "round_submitted": True,
        "practice_completed": practice_completed,
        "summary": practice.get("summary", []),
        "quiz": practice.get("quiz", []),
        "request": practice.get("request", {}),
        "mastery": current_mastery,
        "mastery_order": order,
        "error_message": "",
        "stale_reason": "",
    }
    update_practice_state(generation_id, practice_patch)
    practice = get_generation(generation_id)
    return {
        "practice": practice.get("practice", default_practice_state()) if practice else default_practice_state(),
        "results": graded["results"],
        "mastery": graded["mastery"],
        "recommendation": graded["recommendation"],
        "subtopic_to_revise": graded["subtopic_to_revise"],
        "recommendations": graded["recommendations"],
    }


@app.post("/api/student/{generation_id}/check")
async def api_student_check(request: Request, generation_id: str, payload: dict[str, Any]):
    user_id = ensure_guest_user(request)
    generation = get_generation(generation_id)
    if not generation or generation.get("status") != "completed":
        raise HTTPException(status_code=404, detail="Материал недоступен.")

    existing_attempt = load_student_attempt(generation_id, user_id)
    if existing_attempt:
        return existing_attempt

    quiz = generation.get("quiz", [])
    quiz_subtopics_list = quiz_subtopics(quiz)
    answers = payload.get("answers", [])
    if not isinstance(answers, list):
        raise HTTPException(status_code=400, detail="Некорректный формат ответов.")

    answers_by_id = {}
    for item in answers:
        qid = str(item.get("question_id", ""))
        if qid:
            answers_by_id[qid] = item

    results = []
    open_payload = []
    open_subtopics_by_id: dict[str, str] = {}
    for idx, q in enumerate(quiz):
        qid = str(q.get("question_id", idx + 1))
        qtype = q.get("question_type", "multiple_choice")
        subtopic = (q.get("subtopic") or f"Подтема {idx + 1}").strip()
        user_answer = answers_by_id.get(qid, {})
        if not isinstance(user_answer, dict):
            user_answer = {"answer": user_answer}
        if qtype in ("open_ended", "open_question"):
            open_subtopics_by_id[qid] = subtopic
            open_payload.append(
                {
                    "question_id": qid,
                    "question_text": q.get("question_text", ""),
                    "correct_answer": q.get("correct_answer", ""),
                    "student_answer": user_answer.get("student_answer") if isinstance(user_answer.get("student_answer"), str) else str(user_answer.get("answer") or ""),
                }
            )
            continue

        if "is_correct" in user_answer:
            score = 1 if user_answer.get("is_correct") is True else 0
        else:
            try:
                ua = int(user_answer.get("answer"))
            except (TypeError, ValueError):
                ua = -1
            score = 1 if ua == int(q.get("correct_answer", -999)) else 0
        results.append({"question_id": qid, "subtopic": subtopic, "score": score})

    if open_payload:
        if not ML_API_KEY:
            raise HTTPException(status_code=500, detail="Сервис проверки не настроен.")
        ml_client = MLServiceClient(api_key=ML_API_KEY, base_url=ML_URL)
        try:
            graded = await ml_client.grade_open_answers(open_payload)
            for row in graded.get("scores", []):
                qid = str(row.get("question_id", ""))
                results.append(
                    {
                        "question_id": qid,
                        "subtopic": open_subtopics_by_id.get(qid, ""),
                        "score": int(row.get("score", 0)),
                    }
                )
        except MLServiceError as exc:
            raise HTTPException(status_code=500, detail=exc.user_message)

    mastery = build_mastery_from_results(results, quiz_subtopics_list)
    recommendations = build_recommendations_from_mastery(mastery)
    recommendation = summarize_recommendations(recommendations)
    subtopic_to_revise = choose_subtopic_to_revise(recommendations)

    attempt = save_student_attempt(generation_id, user_id, answers, results, recommendation, subtopic_to_revise)

    return {
        "results": results,
        "mastery": attempt["mastery"],
        "recommendation": recommendation,
        "subtopic_to_revise": subtopic_to_revise,
        "recommendations": recommendations,
    }


@app.get("/")
async def root_redirect():
    return FileResponse(PUBLIC_DIR / "teacher" / "index.html")


@app.get("/teacher/")
@app.get("/teacher/index.html")
async def teacher_redirect():
    return RedirectResponse(url="/")


@app.get("/student/index.html")
async def legacy_student_redirect(request: Request):
    generation_id = request.query_params.get("generation_id", "").strip()
    if generation_id:
        return RedirectResponse(url=f"/material/{generation_id}/")
    return FileResponse(PUBLIC_DIR / "student" / "index.html")


@app.get("/material/{generation_id}")
async def material_redirect(generation_id: str):
    return FileResponse(PUBLIC_DIR / "student" / "index.html")


@app.get("/material/{generation_id}/")
async def material_page(generation_id: str):
    return FileResponse(PUBLIC_DIR / "student" / "index.html")


@app.websocket("/ws/generations")
async def websocket_generations(websocket: WebSocket):
    user_id = websocket.query_params.get("user_id")
    if not user_id:
        await websocket.close(code=1008)
        return
    await websocket.accept()
    WS_CONNECTIONS.setdefault(user_id, set()).add(websocket)
    try:
        await websocket.send_text(json.dumps({"type": "connected"}, ensure_ascii=False))
        while True:
            _ = await websocket.receive_text()
            await websocket.send_text(json.dumps({"type": "pong"}, ensure_ascii=False))
    except WebSocketDisconnect:
        pass
    finally:
        WS_CONNECTIONS.get(user_id, set()).discard(websocket)


app.mount("/", StaticFiles(directory=str(PUBLIC_DIR), html=True), name="static")
